#!/usr/bin/env python3
# Metro Tax Lookup - Arapahoe County
# Copyright (C) 2026 Jesse Lind
# SPDX-License-Identifier: AGPL-3.0-or-later
# See LICENSE for full terms or https://www.gnu.org/licenses/agpl-3.0.html

"""Parcel-record shards + sibling mart / GIS enrich for the new ingest.

Uses logical field names after mapping alias resolve. Arapahoe CSV headers live
only in tools/ingest/mappings/arapahoe.json. Does not import the old rebuild script.
"""

from __future__ import annotations

import csv
import json
import math
import re
import statistics
import sys
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any

from ingest.reader import logical_row_from_csv, resolve_role_column_map
from ingest.situs import accumulate_situs_row, finalize_situs_map
from ingest.out_dir_policy import validate_out_dir

# Keep in sync with PARCEL_RECORD_SHARD_PREFIX_LENGTH in arapahoeParcelLevyData.ts
PARCEL_RECORD_SHARD_PREFIX_LEN = 6
GIS_PARCELS_DATA_AS_OF_FILENAME = "data-as-of.txt"
_DEFAULT_GIS_LAYER = "Assessor_Parcels"
_DEFAULT_GIS_PIN_COL = "PIN"
_DEFAULT_GIS_CODE_COL = "Neighborhood_Code"
_DEFAULT_GIS_NAME_COL = "Neighborhood"

# Colorado DPT residential assessed rates (AssessmentYear >= 2025).
COLORADO_SCHOOL_ASSESSED_RATE = 0.0705
COLORADO_LOCAL_ASSESSED_RATE = 0.068
DUAL_ASSESSED_MIN_ASSESSMENT_YEAR = 2025

_LEGAL_DESCR_PREFIX_RE = re.compile(
    r"^SubdivisionCd\s+\S+\s+SubdivisionName\s+.+?\s+Block\s+\S+\s+Lot\s+\S+\s+",
    re.IGNORECASE,
)
_LEGAL_DESCR_TYPE_RANK = (
    "MetesBounds",
    "UnPlatted",
    "PersonalDescr",
    "Platted",
)

# County PPINum attribute labels → logical building fields (mapping aliases).
BUILDING_ATTRIBUTE_FIELDS: tuple[tuple[str, str | None], ...] = (
    ("Quality Grade", "quality_cd"),
    ("Improvement Type", "impr_tp_dscr"),
    ("Bedrooms", "bed_count"),
    ("Bathrooms", "bath_count"),
    ("Architectural", "impr_mdl_cd_dscr"),
    ("Heat Method", "heat_cd1"),
    ("Cool Method", "cool_cd1"),
    ("Year Built", "act_year"),
    ("Roof", "roof_cd1"),
    ("Fireplaces", "fireplaces"),
    ("Exterior Wall", "extwall_cd1"),
    ("Construction Type", "class"),
)
COUNTY_DECIMAL_ATTRIBUTE_KEYS = frozenset({"bed_count", "bath_count"})
# Arapahoe PPINum shows an empty Fireplaces row when the mart has no value.
ALWAYS_EMIT_BUILDING_ATTRIBUTE_LABELS = frozenset({"Fireplaces"})

PERMIT_STATUS_LABELS: dict[str, str] = {
    "P": "Pending",
    "C": "Complete",
    "A": "Assigned",
    "S": "Submitted",
    "V": "Void",
    "M": "Incomplete",
}


def _strip(val: Any) -> str:
    if val is None:
        return ""
    return str(val).strip()


def _sequence_sort_key(raw: Any) -> tuple[int, float | str]:
    """Sort land/building sequence numbers without raising on odd CSV cells.

    Numeric values (including "1.0") sort together. Non-numeric (e.g. "1A")
    sort after numbers by string. Empty sorts as 0.
    """
    s = _strip(raw)
    if not s:
        return (0, 0.0)
    try:
        return (0, float(s))
    except ValueError:
        return (1, s)


def _optional_str(val: Any) -> str | None:
    s = _strip(val)
    return s if s else None


def normalize_pin(raw: str, pin_digits: int = 9) -> str:
    """Normalize account id for shard keys.

    Digit-only ids are zero-padded (Arapahoe). Letter+digit ids (Douglas) are
    uppercased with spaces/dashes stripped — digits-only strip would destroy
    path-safe prefixes like ``R01039``.
    """
    s = _strip(raw)
    if not s:
        return ""
    m = re.fullmatch(r"(-?\d+)\.0+", s)
    if m:
        s = m.group(1)
    if s.isdigit():
        return s.zfill(pin_digits)[:pin_digits]
    cleaned = re.sub(r"[\s-]", "", s).upper()
    if re.fullmatch(r"[A-Z0-9]+", cleaned):
        return cleaned
    digits = re.sub(r"\D", "", s)
    if not digits:
        return ""
    return digits.zfill(pin_digits)[:pin_digits]


def parse_parcel_value_cell(val: Any) -> float | None:
    if val is None:
        return None
    if isinstance(val, (int, float)):
        if isinstance(val, float) and math.isnan(val):
            return None
        return float(val)
    s = _strip(str(val))
    if not s:
        return None
    try:
        return float(s.replace(",", ""))
    except ValueError:
        return None


def normalize_integerish_code(raw: str | None) -> str:
    s = _strip(raw)
    if not s:
        return ""
    m = re.fullmatch(r"(-?\d+)\.0+", s)
    if m:
        return m.group(1)
    return s


def normalize_state_use_cd(raw: str) -> str:
    s = normalize_integerish_code(raw)
    if s.isdigit() and len(s) == 3:
        return s.zfill(4)
    return s


def parse_assessment_year_cell(val: Any) -> int | None:
    s = _strip(str(val)) if val is not None else ""
    if not s:
        return None
    try:
        year = int(float(s))
    except ValueError:
        return None
    return year if 1900 <= year <= 2100 else None


def is_residential_state_use_code(state_use_cd: str | None) -> bool:
    code = normalize_integerish_code(state_use_cd or "")
    return bool(code) and code.startswith("1")


def parcel_row_qualifies_for_dual_assessed_splits(row: dict[str, str]) -> bool:
    year = parse_assessment_year_cell(row.get("assessment_year"))
    if year is None or year < DUAL_ASSESSED_MIN_ASSESSMENT_YEAR:
        return False
    return _strip(row.get("tax_roll_descr", "")).upper() == "REAL"


def parcel_row_qualifies_for_school_assessed_splits(row: dict[str, str]) -> bool:
    if not parcel_row_qualifies_for_dual_assessed_splits(row):
        return False
    if not is_residential_state_use_code(row.get("state_use_cd")):
        return False
    return _strip(row.get("property_class_descr", "")) == "Improvement"


def _positive_actual(val: float | None) -> float:
    if val is None or not math.isfinite(val) or val <= 0:
        return 0.0
    return float(val)


def round_school_assessed_component(actual: float | None) -> int | None:
    if actual is None or not math.isfinite(actual):
        return None
    return round(actual * COLORADO_SCHOOL_ASSESSED_RATE)


def school_assessed_fields_from_actuals(
    improvement_actual: float | None,
    land_actual: float | None,
    total_actual: float | None,
) -> dict[str, int]:
    building = round_school_assessed_component(improvement_actual)
    land = round_school_assessed_component(land_actual)
    if building is not None or land is not None:
        out: dict[str, int] = {}
        if building is not None:
            out["schoolAssessedBuilding"] = building
        if land is not None:
            out["schoolAssessedLand"] = land
        out["schoolAssessedTotal"] = (building or 0) + (land or 0)
        return out
    total = round_school_assessed_component(total_actual)
    if total is not None:
        return {"schoolAssessedTotal": total}
    return {}


def local_assessed_split_fields(
    improvement_actual: float | None,
    land_actual: float | None,
    total_assessed: float | None,
) -> dict[str, int]:
    if total_assessed is None or not math.isfinite(total_assessed):
        return {}
    total_int = round(total_assessed)
    imp = _positive_actual(improvement_actual)
    land = _positive_actual(land_actual)
    if imp == 0 and land > 0:
        return {"assessedBuilding": 0, "assessedLand": total_int}
    if imp > 0 and land == 0:
        return {"assessedBuilding": total_int, "assessedLand": 0}
    if imp > 0 and land > 0:
        land_assessed = round(land * COLORADO_LOCAL_ASSESSED_RATE)
        return {
            "assessedLand": land_assessed,
            "assessedBuilding": max(0, total_int - land_assessed),
        }
    return {}


def non_residential_assessed_split_fields(
    improvement_actual: float | None,
    land_actual: float | None,
    total_actual: float | None,
    total_assessed: float | None,
) -> dict[str, int]:
    if total_assessed is None or not math.isfinite(total_assessed):
        return {}
    total_int = round(total_assessed)
    imp = _positive_actual(improvement_actual)
    land = _positive_actual(land_actual)
    actual_total = _positive_actual(total_actual)
    if imp == 0 and land > 0:
        return {"assessedBuilding": 0, "assessedLand": total_int}
    if imp > 0 and land == 0:
        return {"assessedBuilding": total_int, "assessedLand": 0}
    if imp > 0 and land > 0 and actual_total > 0:
        building_assessed = round(total_int * (imp / actual_total))
        return {
            "assessedBuilding": building_assessed,
            "assessedLand": max(0, total_int - building_assessed),
        }
    return {}


def attach_computed_assessed_values(rec: dict[str, Any], row: dict[str, str]) -> None:
    if not parcel_row_qualifies_for_dual_assessed_splits(row):
        return
    if is_residential_state_use_code(row.get("state_use_cd")):
        rec.update(
            local_assessed_split_fields(
                rec.get("improvementActual"),
                rec.get("landActual"),
                rec.get("totalAssessed"),
            )
        )
    else:
        rec.update(
            non_residential_assessed_split_fields(
                rec.get("improvementActual"),
                rec.get("landActual"),
                rec.get("totalActual"),
                rec.get("totalAssessed"),
            )
        )
    if parcel_row_qualifies_for_school_assessed_splits(row):
        rec.update(
            school_assessed_fields_from_actuals(
                rec.get("improvementActual"),
                rec.get("landActual"),
                rec.get("totalActual"),
            )
        )


def legal_descr_display_tail(full: str) -> str:
    s = _strip(full)
    if not s:
        return ""
    m = _LEGAL_DESCR_PREFIX_RE.match(s)
    if m:
        tail = s[m.end() :].strip()
        return tail if tail else s
    return s


def _situs_address_from_logical(row: dict[str, str]) -> str | None:
    """Prefer free-form situs; otherwise compose from street parts (Douglas)."""
    free = _optional_str(row.get("sa_free_form_addr"))
    if free:
        return free
    from ingest.situs import format_situs_label

    label = format_situs_label(row)
    pin = _strip(row.get("pin", ""))
    if not label or label == pin:
        return None
    # Drop locality suffix for the Property details situs line (city is separate).
    city = _strip(row.get("sa_city", ""))
    if city and f", {city}" in label:
        return label.split(f", {city}", 1)[0].strip() or None
    return label


def _owner_city_state_zip_from_logical(row: dict[str, str]) -> str | None:
    composed = _optional_str(row.get("cur_last_line"))
    if composed:
        return composed
    city = _strip(row.get("mailing_city", ""))
    state = _strip(row.get("mailing_state", ""))
    zipc = _strip(row.get("mailing_zip", ""))
    if zipc and len(zipc) > 5 and zipc[:5].isdigit():
        zipc = f"{zipc[:5]}-{zipc[5:]}" if len(zipc) >= 9 else zipc[:5]
    locality = " ".join(x for x in (city, state) if x)
    if locality and zipc:
        return f"{locality} {zipc}"
    return locality or zipc or None


def _owner_delivery_from_logical(row: dict[str, str]) -> str | None:
    line1 = _strip(row.get("cur_delivery_addr", ""))
    line2 = _strip(row.get("mailing_address_line_2", ""))
    joined = ", ".join(x for x in (line1, line2) if x)
    return joined or None


def parcel_record_from_logical_row(row: dict[str, str]) -> dict[str, Any]:
    """Build one parcel-record entry from a logical Main Parcel / location row."""
    legal_full = _optional_str(row.get("legal_descr"))
    legal_display = legal_descr_display_tail(legal_full) if legal_full else None
    rec: dict[str, Any] = {
        "ain": _optional_str(row.get("ain")),
        "situsAddress": _situs_address_from_logical(row),
        "situsCity": _optional_str(row.get("sa_city")),
        "ownerList": _optional_str(row.get("owner_list")),
        "ownerDeliveryAddress": _owner_delivery_from_logical(row),
        "ownerCityStateZip": _owner_city_state_zip_from_logical(row),
        "legalDescrFull": legal_full,
        "legalDescrDisplay": legal_display,
        "subdivisionCd": normalize_integerish_code(row.get("subdivision_cd")) or None,
        "subdivisionName": _optional_str(row.get("subdivision_name")),
        "taxRollDescr": _optional_str(row.get("tax_roll_descr")),
        "propertyClassDescr": _optional_str(row.get("property_class_descr")),
        "totalActual": parse_parcel_value_cell(row.get("total_actual")),
        "improvementActual": parse_parcel_value_cell(row.get("improvement_actual")),
        "landActual": parse_parcel_value_cell(row.get("land_actual")),
        "totalAssessed": parse_parcel_value_cell(row.get("total_assessed")),
        "stateUseCd": normalize_integerish_code(row.get("state_use_cd")) or None,
        "parcelTaxYear": _optional_str(row.get("parcel_tax_year")),
        "assessmentYear": _optional_str(row.get("assessment_year")),
    }
    nbhd_code = format_neighborhood_code_with_extension(
        row.get("neighborhood_code"),
        row.get("neighborhood_extention"),
    )
    if nbhd_code:
        rec["neighborhoodCode"] = nbhd_code
    acre = parse_parcel_value_cell(row.get("acreage"))
    if acre is not None:
        rec["acreage"] = format_acreage_display(acre)
    attach_computed_assessed_values(rec, row)
    return rec


def account_row_from_logical(row: dict[str, str]) -> dict[str, Any] | None:
    """Intermediate account-map row (same shape as reader.read_account_rows)."""
    account_id = _strip(row.get("pin", ""))
    if not account_id:
        return None
    return {
        "accountId": account_id,
        "taxAreaId": _strip(row.get("tag_id", "")),
        "tagShortDescr": _optional_str(row.get("tag_short_descr")),
        "totalActual": parse_parcel_value_cell(row.get("total_actual")),
        "totalAssessed": parse_parcel_value_cell(row.get("total_assessed")),
        "parcelTaxYear": _optional_str(row.get("parcel_tax_year")),
        "assessmentYear": _optional_str(row.get("assessment_year")),
        "propertyClassDescr": _optional_str(row.get("property_class_descr")),
        "ownerList": _optional_str(row.get("owner_list")),
        "ain": _optional_str(row.get("ain")),
    }


def read_main_parcel_bundle(
    csv_path: Path,
    mapping: dict[str, Any],
) -> tuple[list[dict[str, Any]], dict[str, list[dict[str, str]]], dict[str, dict[str, Any]]]:
    """
    One Main Parcel pass: account intermediate rows, situs map, parcel-record map.

    Account rows mirror read_account_rows (one entry per CSV row with a pin).
    """
    pin_digits = int(mapping.get("identifierDigits", 9))
    account_rows: list[dict[str, Any]] = []
    situs_by_key: dict[str, dict[str, str]] = {}
    parcel_record_map: dict[str, dict[str, Any]] = {}

    with csv_path.open(newline="", encoding="utf-8-sig", errors="replace") as f:
        reader = csv.DictReader(f)
        headers = list(reader.fieldnames or [])
        col_map = resolve_role_column_map(headers, mapping, "parcel")
        if "pin" not in col_map:
            raise ValueError(f"{csv_path}: pin column not found via mapping aliases")

        for raw in reader:
            logical = logical_row_from_csv(raw, col_map)
            pin_raw = logical.get("pin", "")
            pin = normalize_pin(pin_raw, pin_digits)
            if not pin:
                continue
            # Keep raw pin string for account map normalize (writer zero-pads).
            acct = account_row_from_logical({**logical, "pin": pin_raw})
            if acct:
                account_rows.append(acct)

            accumulate_situs_row(situs_by_key, {**logical, "pin": pin}, pin)

            if pin not in parcel_record_map:
                parcel_record_map[pin] = parcel_record_from_logical_row(logical)
            else:
                ain = _optional_str(logical.get("ain"))
                if ain and not parcel_record_map[pin].get("ain"):
                    parcel_record_map[pin]["ain"] = ain

    return account_rows, finalize_situs_map(situs_by_key), parcel_record_map


def read_location_parcel_record_map(
    location_path: Path,
    mapping: dict[str, Any],
    *,
    values_totals: dict[str, dict[str, float]] | None = None,
) -> dict[str, dict[str, Any]]:
    """
    Build parcel-record rows from a location/account file (Douglas valuesFile path).

    Uses accountMap.file role aliases. Optional values_totals overlays actual/assessed.
    """
    from ingest.reader import open_csv_dict_reader

    acct_cfg = mapping["accountMap"]
    file_role = acct_cfg["file"]
    pin_digits = int(mapping.get("identifierDigits", 9))
    account_alias = acct_cfg.get("accountId", "")
    parcel_record_map: dict[str, dict[str, Any]] = {}

    with open_csv_dict_reader(location_path, mapping, file_role) as (reader, headers):
        col_map = resolve_role_column_map(headers, mapping, file_role)
        pin_alias = "pin" if "pin" in col_map else account_alias
        if pin_alias not in col_map:
            raise ValueError(
                f"{location_path}: parcel-record build needs pin/account column via mapping"
            )
        for raw in reader:
            logical = logical_row_from_csv(raw, col_map)
            pin_raw = logical.get(pin_alias, "") or logical.get(account_alias, "")
            pin = normalize_pin(pin_raw, pin_digits)
            if not pin:
                continue
            if pin not in parcel_record_map:
                parcel_record_map[pin] = parcel_record_from_logical_row(
                    {**logical, "pin": pin}
                )
            else:
                ain = _optional_str(logical.get("ain"))
                if ain and not parcel_record_map[pin].get("ain"):
                    parcel_record_map[pin]["ain"] = ain

    if values_totals:
        for pin, rec in parcel_record_map.items():
            bucket = values_totals.get(pin)
            if not bucket:
                continue
            if rec.get("totalActual") is None:
                rec["totalActual"] = bucket.get("totalActual")
            if rec.get("totalAssessed") is None:
                rec["totalAssessed"] = bucket.get("totalAssessed")

    return parcel_record_map


# -----------------------------------------------------------------------
# Sibling mart readers (logical fields)
# -----------------------------------------------------------------------

def _read_logical_csv_rows(
    path: Path,
    mapping: dict[str, Any],
    file_role: str,
) -> list[dict[str, str]]:
    if not path.is_file():
        return []
    from ingest.reader import open_csv_dict_reader

    with open_csv_dict_reader(path, mapping, file_role) as (reader, headers):
        col_map = resolve_role_column_map(headers, mapping, file_role)
        return [logical_row_from_csv(raw, col_map) for raw in reader]


def _legal_descr_type_rank(descr_type: str) -> int:
    t = _strip(descr_type)
    try:
        return _LEGAL_DESCR_TYPE_RANK.index(t)
    except ValueError:
        return len(_LEGAL_DESCR_TYPE_RANK)


def pick_legal_description_display(rows: list[tuple[str, str]]) -> str | None:
    preferred = frozenset({"MetesBounds", "UnPlatted", "PersonalDescr"})
    sorted_rows = sorted(rows, key=lambda item: _legal_descr_type_rank(item[0]))
    for dtype, display in sorted_rows:
        if dtype not in preferred:
            continue
        tail = legal_descr_display_tail(display)
        if tail:
            return tail
    return None


def read_legal_description_display_by_pin(
    path: Path,
    mapping: dict[str, Any],
    *,
    pin_digits: int = 9,
) -> dict[str, str]:
    by_pin: dict[str, list[tuple[str, str]]] = defaultdict(list)
    for row in _read_logical_csv_rows(path, mapping, "legalDescriptions"):
        pin = normalize_pin(row.get("pin", ""), pin_digits)
        display = _strip(row.get("display_descr", ""))
        dtype = _strip(row.get("descr_type", ""))
        if not pin or not display:
            continue
        by_pin[pin].append((dtype, display))
    out: dict[str, str] = {}
    for pin, items in by_pin.items():
        picked = pick_legal_description_display(items)
        if picked:
            out[pin] = picked
    return out


def ownership_type_label_from_owner_lp_types(lp_types: list[str]) -> str | None:
    types = [_strip(t) for t in lp_types if _strip(t)]
    if not types:
        return None
    if len(types) == 1:
        return types[0]
    unique = list(dict.fromkeys(types))
    if len(unique) == 1 and unique[0] == "Individual":
        return "Joint Tenancy"
    return ", ".join(unique)


def read_ownership_type_by_pin(
    path: Path,
    mapping: dict[str, Any],
    *,
    pin_digits: int = 9,
) -> dict[str, str]:
    by_pin: dict[str, list[str]] = defaultdict(list)
    for row in _read_logical_csv_rows(path, mapping, "legalParties"):
        pin = normalize_pin(row.get("pin", ""), pin_digits)
        if _strip(row.get("lpr_type", "")).upper() != "OWNER":
            continue
        lp_type = _strip(row.get("lp_type", ""))
        if not pin or not lp_type:
            continue
        by_pin[pin].append(lp_type)
    out: dict[str, str] = {}
    for pin, lp_types in by_pin.items():
        label = ownership_type_label_from_owner_lp_types(lp_types)
        if label:
            out[pin] = label
    return out


def read_ownership_name_fields_by_pin(
    path: Path,
    mapping: dict[str, Any],
    *,
    pin_digits: int = 9,
) -> dict[str, dict[str, str]]:
    """Owner name + mailing from Assessor ownership tables (Douglas)."""
    out: dict[str, dict[str, str]] = {}
    for row in _read_logical_csv_rows(path, mapping, "ownership"):
        pin = normalize_pin(row.get("pin", ""), pin_digits)
        if not pin or pin in out:
            continue
        entry: dict[str, str] = {}
        owner = _optional_str(row.get("owner_list"))
        if owner:
            entry["ownerList"] = owner
        delivery = _owner_delivery_from_logical(row)
        if delivery:
            entry["ownerDeliveryAddress"] = delivery
        city_state_zip = _owner_city_state_zip_from_logical(row)
        if city_state_zip:
            entry["ownerCityStateZip"] = city_state_zip
        if entry:
            out[pin] = entry
    return out


def apply_ownership_names_to_account_rows(
    account_rows: list[dict[str, Any]],
    ownership_by_pin: dict[str, dict[str, str]],
    *,
    pin_digits: int,
) -> list[dict[str, Any]]:
    """Fill missing account-map ownerList from ownership join (summary tiles)."""
    if not ownership_by_pin:
        return account_rows
    out: list[dict[str, Any]] = []
    for row in account_rows:
        merged = dict(row)
        if merged.get("ownerList"):
            out.append(merged)
            continue
        account_id = normalize_pin(str(merged.get("accountId", "")), pin_digits)
        fields = ownership_by_pin.get(account_id) if account_id else None
        owner = fields.get("ownerList") if fields else None
        if owner:
            merged["ownerList"] = owner
        out.append(merged)
    return out


def read_subdivision_fields_by_pin(
    path: Path,
    mapping: dict[str, Any],
    *,
    pin_digits: int = 9,
) -> dict[str, dict[str, str]]:
    """Subdivision name / filing code from Assessor subdivision table (Douglas)."""
    out: dict[str, dict[str, str]] = {}
    for row in _read_logical_csv_rows(path, mapping, "subdivision"):
        pin = normalize_pin(row.get("pin", ""), pin_digits)
        if not pin or pin in out:
            continue
        entry: dict[str, str] = {}
        name = _optional_str(row.get("subdivision_name"))
        code = normalize_integerish_code(row.get("subdivision_cd")) or None
        lot = _optional_str(row.get("lot_no"))
        block = _optional_str(row.get("block_no"))
        tract = _optional_str(row.get("tract_no"))
        if name:
            entry["subdivisionName"] = name
        if code:
            entry["subdivisionCd"] = code
        if lot and lot != "0":
            entry["lotNo"] = lot
        if block and block != "0":
            entry["blockNo"] = block
        if tract and tract != "0":
            entry["tractNo"] = tract
        if entry:
            out[pin] = entry
    return out


def format_acreage_display(total: float) -> str:
    if not math.isfinite(total):
        return ""
    return f"{total:.4f}"


def format_land_units_display(row: dict[str, str]) -> str:
    uts = _strip(row.get("uts", ""))
    unit_tp = _strip(row.get("unit_tp", ""))
    if not uts:
        return ""
    try:
        units = f"{float(uts):.4f}"
    except ValueError:
        units = uts
    return f"{units} {unit_tp}".strip() if unit_tp else units


def land_table_row_from_logical(row: dict[str, str]) -> dict[str, str] | None:
    units = format_land_units_display(row)
    land_use = _strip(row.get("use_cd_dscr", ""))
    if not units and not land_use:
        return None
    out: dict[str, str] = {}
    if units:
        out["units"] = units
    if land_use:
        out["landUse"] = land_use
    return out


def read_land_fields_by_pin(
    path: Path,
    mapping: dict[str, Any],
    *,
    pin_digits: int = 9,
) -> dict[str, dict[str, Any]]:
    by_pin: dict[str, list[dict[str, str]]] = defaultdict(list)
    for row in _read_logical_csv_rows(path, mapping, "land"):
        pin = normalize_pin(row.get("pin", ""), pin_digits)
        if pin:
            by_pin[pin].append(row)
    out: dict[str, dict[str, Any]] = {}
    for pin, rows in by_pin.items():
        rows_sorted = sorted(
            rows,
            key=lambda r: _sequence_sort_key(r.get("num", "")),
        )
        land_lines = [
            line
            for line in (land_table_row_from_logical(r) for r in rows_sorted)
            if line
        ]
        total_acre = 0.0
        has_acre = False
        for row in rows_sorted:
            acre = parse_parcel_value_cell(row.get("acreage"))
            if acre is not None:
                total_acre += acre
                has_acre = True
        entry: dict[str, Any] = {}
        if has_acre:
            entry["acreage"] = format_acreage_display(total_acre)
        if land_lines:
            entry["landLines"] = land_lines
        if entry:
            out[pin] = entry
    return out


def format_county_count(val: Any) -> str:
    s = _strip(str(val)) if val is not None else ""
    if not s:
        return ""
    try:
        return f"{float(s):.2f}"
    except ValueError:
        return s


def format_county_sqft(val: Any) -> str:
    s = _strip(str(val)) if val is not None else ""
    if not s:
        return ""
    try:
        n = float(s)
        if n == int(n):
            return str(int(n))
        return s
    except ValueError:
        return s


def building_record_from_logical(row: dict[str, str]) -> dict[str, Any] | None:
    building_num = _strip(row.get("num", ""))
    attributes: list[dict[str, str]] = []
    for label, key in BUILDING_ATTRIBUTE_FIELDS:
        if key is None:
            if label in ALWAYS_EMIT_BUILDING_ATTRIBUTE_LABELS:
                attributes.append({"label": label, "value": ""})
            continue
        raw = _strip(row.get(key, ""))
        if not raw:
            if label in ALWAYS_EMIT_BUILDING_ATTRIBUTE_LABELS:
                attributes.append({"label": label, "value": ""})
            continue
        if key in COUNTY_DECIMAL_ATTRIBUTE_KEYS:
            value = format_county_count(raw)
        else:
            value = raw
        attributes.append({"label": label, "value": value})
    areas: list[dict[str, str]] = []
    for i in range(1, 20):
        descr = _strip(row.get(f"sar_cat_dscr_{i}", ""))
        if not descr:
            continue
        area_raw = _strip(row.get(f"sar_cat_area_{i}", ""))
        areas.append(
            {
                "description": descr,
                "sqFt": format_county_sqft(area_raw) if area_raw else "",
            }
        )
    total_area = format_county_sqft(row.get("base_area"))
    if total_area and not areas:
        areas.append({"description": "Improvement", "sqFt": total_area})
    if not building_num and not attributes and not areas:
        return None
    rec: dict[str, Any] = {"buildingNum": building_num or "1"}
    if attributes:
        rec["attributes"] = attributes
    if areas:
        rec["areas"] = areas
    if total_area:
        rec["totalArea"] = total_area
    return rec


def read_building_fields_by_pin(
    bld_path: Path,
    mapping: dict[str, Any],
    *,
    pin_digits: int = 9,
) -> dict[str, dict[str, Any]]:
    bld_by_pin: dict[str, list[dict[str, str]]] = defaultdict(list)
    for row in _read_logical_csv_rows(bld_path, mapping, "building"):
        pin = normalize_pin(row.get("pin", ""), pin_digits)
        if pin:
            bld_by_pin[pin].append(row)
    out: dict[str, dict[str, Any]] = {}
    for pin, rows in bld_by_pin.items():
        bld_rows = sorted(
            rows,
            key=lambda r: _sequence_sort_key(r.get("num", "")),
        )
        buildings = [
            bld for bld in (building_record_from_logical(r) for r in bld_rows) if bld
        ]
        entry: dict[str, Any] = {}
        for row in bld_rows:
            typ = _strip(row.get("impr_tp_dscr", ""))
            if typ:
                entry["landUse"] = typ
                break
        if buildings:
            entry["buildings"] = buildings
        if entry:
            out[pin] = entry
    return out


def format_county_mm_dd_yyyy(raw: str) -> str:
    s = _strip(raw)
    if len(s) == 8 and s.isdigit():
        return f"{s[4:6]}-{s[6:8]}-{s[0:4]}"
    return s


def format_sale_date_display(raw: str) -> tuple[str, str]:
    """Return (display MM-DD-YYYY, sort key YYYYMMDD or raw)."""
    s = _strip(raw)
    if not s:
        return "", ""
    m = re.match(r"(\d{4})-(\d{2})-(\d{2})", s)
    if m:
        y, mo, d = m.group(1), m.group(2), m.group(3)
        return f"{mo}-{d}-{y}", f"{y}{mo}{d}"
    m = re.match(r"(\d{1,2})/(\d{1,2})/(\d{2,4})", s)
    if m:
        mo, d, y = m.group(1), m.group(2), m.group(3)
        if len(y) == 2:
            y = f"20{y}" if int(y) < 70 else f"19{y}"
        return f"{int(mo):02d}-{int(d):02d}-{y}", f"{y}{int(mo):02d}{int(d):02d}"
    display = format_county_mm_dd_yyyy(s)
    sort_key = s if len(s) == 8 and s.isdigit() else s
    return display, sort_key


def format_book_page_display(book: str, page: str) -> str:
    b = _strip(book)
    p = _strip(page)
    if not b and not p:
        return ""
    return f"{b} {p}".strip()


def transfer_sale_row_from_logical(row: dict[str, str]) -> dict[str, Any] | None:
    book = _strip(row.get("book", ""))
    page = _strip(row.get("page", ""))
    recording = _strip(row.get("recording_no", ""))
    book_page = format_book_page_display(book, page) or recording
    date, sort_date = format_sale_date_display(row.get("doc_date", ""))
    price = parse_parcel_value_cell(row.get("consid"))
    deed_type = _optional_str(row.get("deed_type"))
    grantor = _optional_str(row.get("grantor"))
    grantee = _optional_str(row.get("grantee"))
    if not book_page and not date and price is None:
        return None
    out: dict[str, Any] = {
        "bookPage": book_page,
        "date": date or "",
        "sortDate": sort_date or _strip(row.get("doc_date", "")),
    }
    if price is not None:
        out["price"] = price
    if deed_type:
        out["type"] = deed_type
    if grantor:
        out["grantor"] = grantor
    if grantee:
        out["grantee"] = grantee
    return out


def read_transfers_by_pin(
    path: Path,
    mapping: dict[str, Any],
    *,
    pin_digits: int = 9,
) -> dict[str, list[dict[str, Any]]]:
    by_pin: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in _read_logical_csv_rows(path, mapping, "transfers"):
        pin = normalize_pin(row.get("pin", ""), pin_digits)
        if not pin:
            continue
        sale = transfer_sale_row_from_logical(row)
        if sale:
            by_pin[pin].append(sale)
    out: dict[str, list[dict[str, Any]]] = {}
    for pin, rows in by_pin.items():
        rows_sorted = sorted(
            rows,
            key=lambda r: _strip(str(r.get("sortDate", ""))),
            reverse=True,
        )
        out[pin] = [{k: v for k, v in r.items() if k != "sortDate"} for r in rows_sorted]
    return out


def permit_row_from_logical(row: dict[str, str]) -> dict[str, Any] | None:
    permit_num = _strip(row.get("permit_num", ""))
    dscr = _strip(row.get("dscr", ""))
    status_raw = _strip(row.get("status", ""))
    issue_raw = _strip(row.get("issue_dt", ""))
    final_raw = _strip(row.get("final_dt", ""))
    est_val = parse_parcel_value_cell(row.get("est_val"))
    if not permit_num and not dscr and est_val is None:
        return None
    status = PERMIT_STATUS_LABELS.get(status_raw.upper(), status_raw)
    issue = (
        format_county_mm_dd_yyyy(issue_raw)
        if issue_raw and issue_raw != "18991230"
        else ""
    )
    final = (
        format_county_mm_dd_yyyy(final_raw)
        if final_raw and final_raw != "18991230"
        else ""
    )
    out: dict[str, Any] = {
        "sortDate": issue_raw if issue_raw and issue_raw != "18991230" else "",
    }
    if permit_num:
        out["permitNum"] = permit_num
    if status:
        out["status"] = status
    if dscr:
        out["description"] = dscr
    if issue:
        out["issueDate"] = issue
    if final:
        out["finalDate"] = final
    if est_val is not None:
        out["estimatedValue"] = est_val
    return out


def read_permits_by_pin(
    path: Path,
    mapping: dict[str, Any],
    *,
    pin_digits: int = 9,
) -> dict[str, list[dict[str, Any]]]:
    by_pin: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in _read_logical_csv_rows(path, mapping, "permits"):
        pin = normalize_pin(row.get("pin", ""), pin_digits)
        if not pin:
            continue
        permit = permit_row_from_logical(row)
        if permit:
            by_pin[pin].append(permit)
    out: dict[str, list[dict[str, Any]]] = {}
    for pin, rows in by_pin.items():
        rows_sorted = sorted(
            rows,
            key=lambda r: (
                _strip(str(r.get("sortDate", ""))),
                _strip(str(r.get("permitNum", ""))),
            ),
            reverse=True,
        )
        out[pin] = [{k: v for k, v in r.items() if k != "sortDate"} for r in rows_sorted]
    return out


def read_xlsx_code_description_map(
    path: Path,
    *,
    code_col: int = 0,
    desc_col: int = 1,
) -> dict[str, str]:
    if not path.is_file():
        return {}
    try:
        from openpyxl import load_workbook
    except ImportError:
        print(
            f"openpyxl not installed; skipping lookup workbook {path.name}",
            file=sys.stderr,
        )
        return {}
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        out: dict[str, str] = {}
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if not row or len(row) <= max(code_col, desc_col):
                continue
            if i == 0:
                continue
            code_raw = row[code_col]
            desc_raw = row[desc_col]
            if code_raw is None:
                continue
            if isinstance(code_raw, (int, float)):
                code = (
                    str(int(code_raw))
                    if float(code_raw) == int(code_raw)
                    else str(code_raw)
                )
            else:
                code = normalize_state_use_cd(str(code_raw))
            if not code:
                continue
            desc = _strip(str(desc_raw) if desc_raw is not None else "")
            if desc:
                out[code] = desc
        return out
    finally:
        wb.close()


def read_state_class_description_by_code(path: Path) -> dict[str, str]:
    return read_xlsx_code_description_map(path, code_col=0, desc_col=2)


def read_nbhd_description_by_code(path: Path) -> dict[str, str]:
    return read_xlsx_code_description_map(path)


def format_neighborhood_code(raw: Any) -> str:
    if raw is None:
        return ""
    if isinstance(raw, bool):
        return ""
    if isinstance(raw, int):
        return str(raw) if raw > 0 else ""
    if isinstance(raw, float):
        if math.isnan(raw) or raw <= 0:
            return ""
        if raw == int(raw):
            return str(int(raw))
        return _strip(str(raw))
    s = _strip(str(raw))
    if not s:
        return ""
    if s.lower() == "null":
        return ""
    try:
        val = float(s)
    except ValueError:
        return s
    if math.isnan(val) or val <= 0:
        return ""
    if val == int(val):
        return str(int(val))
    return s


def format_neighborhood_code_with_extension(
    code_raw: Any,
    extension_raw: Any,
) -> str | None:
    """Douglas Assessor location: Neighborhood_Code + Neighborhood_Extention."""
    code = format_neighborhood_code(code_raw)
    if not code:
        return None
    ext = _strip(str(extension_raw) if extension_raw is not None else "")
    if not ext or ext.lower() in {"null", "00"}:
        return code
    return f"{code}-{ext}"


def format_valuation_amount_display(raw: Any) -> str:
    val = parse_parcel_value_cell(raw)
    if val is None:
        return ""
    if val == int(val):
        return f"{int(val):,}"
    return f"{val:,.2f}".rstrip("0").rstrip(".")


def read_values_parcel_enrichment_by_pin(
    path: Path,
    mapping: dict[str, Any],
    *,
    pin_digits: int = 9,
) -> dict[str, dict[str, Any]]:
    """
    Douglas Property_Values.txt: valuation-class rows per account.

    Emits landLines (class description + actual value), land/improvement actual
    splits (Valuation_Type_Code L vs I), and a representative stateUseCd.
    """
    from ingest.reader import open_csv_dict_reader

    acct_cfg = mapping.get("accountMap") or {}
    values_role = acct_cfg.get("valuesFile")
    if not isinstance(values_role, str) or not values_role:
        return {}
    account_alias = acct_cfg.get("accountId", "account_no")
    by_pin: dict[str, list[dict[str, str]]] = defaultdict(list)
    with open_csv_dict_reader(path, mapping, values_role) as (reader, headers):
        col_map = resolve_role_column_map(headers, mapping, values_role)
        if account_alias not in col_map:
            return {}
        for raw in reader:
            logical = logical_row_from_csv(raw, col_map)
            pin = normalize_pin(logical.get(account_alias, ""), pin_digits)
            if pin:
                by_pin[pin].append(logical)
    out: dict[str, dict[str, Any]] = {}
    for pin, rows in by_pin.items():
        land_lines: list[dict[str, str]] = []
        land_actual = 0.0
        improvement_actual = 0.0
        land_assessed = 0.0
        improvement_assessed = 0.0
        state_use_cd: str | None = None
        largest_actual = -1.0
        for row in rows:
            actual = parse_parcel_value_cell(row.get("actual_value"))
            assessed = parse_parcel_value_cell(row.get("assessed_value"))
            descr = _optional_str(row.get("valuation_description")) or ""
            type_code = _strip(row.get("valuation_type_code", "")).upper()
            class_code = normalize_integerish_code(row.get("valuation_class_code"))
            if class_code and actual is not None and actual >= largest_actual:
                largest_actual = actual
                state_use_cd = class_code
            units = format_valuation_amount_display(row.get("actual_value"))
            land_line = land_table_row_from_logical(
                {
                    "uts": units,
                    "unit_tp": "",
                    "use_cd_dscr": descr,
                }
            )
            if land_line:
                land_lines.append(land_line)
            if type_code == "L":
                if actual is not None:
                    land_actual += actual
                if assessed is not None:
                    land_assessed += assessed
            elif type_code == "I":
                if actual is not None:
                    improvement_actual += actual
                if assessed is not None:
                    improvement_assessed += assessed
        entry: dict[str, Any] = {}
        if land_lines:
            entry["landLines"] = land_lines
        if land_actual > 0:
            entry["landActual"] = land_actual
        if improvement_actual > 0:
            entry["improvementActual"] = improvement_actual
        if state_use_cd:
            entry["stateUseCd"] = state_use_cd
        if entry:
            out[pin] = entry
    return out


def read_filing_lookup_by_recording_no(
    path: Path,
    mapping: dict[str, Any],
) -> dict[str, dict[str, str]]:
    """Douglas Property_Filing.txt: recording number → filing description."""
    out: dict[str, dict[str, str]] = {}
    for row in _read_logical_csv_rows(path, mapping, "filing"):
        recording = normalize_integerish_code(row.get("sub_filing_recording_no")) or _strip(
            row.get("sub_filing_recording_no", "")
        )
        if not recording or recording in out:
            continue
        entry: dict[str, str] = {}
        descr = _optional_str(row.get("filing_descr"))
        filing_no = _optional_str(row.get("filing_no"))
        subdivision_name = _optional_str(row.get("subdivision_name"))
        if descr:
            entry["filingDescr"] = descr
        if filing_no:
            entry["filingNo"] = filing_no
        if subdivision_name:
            entry["subdivisionName"] = subdivision_name
        if entry:
            out[recording] = entry
    return out


def _parcels_csv_config(mapping: dict[str, Any]) -> dict[str, str]:
    cfg = (mapping or {}).get("parcelsCsv") or {}
    return {
        "pin": _strip(cfg.get("pin")) or "ACCOUNT_NO",
        "blockNo": _strip(cfg.get("blockNo")) or "BLOCK_NO",
        "tract": _strip(cfg.get("tract")) or "TRACT",
        "filingDescr": _strip(cfg.get("filingDescr")) or "FILING_DESCR",
        "legalDescr": _strip(cfg.get("legalDescr")) or "CAMA_LEGAL_DESC",
    }


def read_parcels_csv_enrichment_by_pin(
    path: Path,
    mapping: dict[str, Any],
    *,
    pin_digits: int = 9,
) -> dict[str, dict[str, str]]:
    """Douglas Hub parcels CSV: block / tract / filing / legal when present."""
    cols = _parcels_csv_config(mapping)
    pin_col = cols["pin"]
    out: dict[str, dict[str, str]] = {}
    with path.open(newline="", encoding="utf-8-sig", errors="replace") as f:
        reader = csv.DictReader(f)
        for raw in reader:
            pin = normalize_pin(raw.get(pin_col, ""), pin_digits)
            if not pin or pin in out:
                continue
            entry: dict[str, str] = {}
            block = _optional_str(raw.get(cols["blockNo"]))
            tract = _optional_str(raw.get(cols["tract"]))
            filing = _optional_str(raw.get(cols["filingDescr"]))
            legal = _optional_str(raw.get(cols["legalDescr"]))
            if block and block != "0":
                entry["blockNo"] = block
            if tract and tract != "0":
                entry["tractNo"] = tract
            if filing:
                entry["filingDescr"] = filing
            if legal:
                entry["legalDescrFull"] = legal
            if entry:
                out[pin] = entry
    return out


def neighborhood_by_pin_from_rows(
    rows: list[tuple[Any, Any, Any]],
    *,
    pin_digits: int = 9,
) -> tuple[dict[str, dict[str, str]], int]:
    provisional: dict[str, tuple[str, str]] = {}
    conflicts: set[str] = set()
    for pin_raw, code_raw, name_raw in rows:
        pin = normalize_pin(str(pin_raw) if pin_raw is not None else "", pin_digits)
        if not pin or pin in conflicts:
            continue
        code = format_neighborhood_code(code_raw)
        if not code:
            continue
        name = _strip(str(name_raw) if name_raw is not None else "")
        if not name:
            continue
        tup = (code, name)
        prev = provisional.get(pin)
        if prev is None:
            provisional[pin] = tup
        elif prev != tup:
            conflicts.add(pin)
            del provisional[pin]
    out = {
        pin: {"neighborhood": name, "neighborhoodCode": code}
        for pin, (code, name) in provisional.items()
    }
    return out, len(conflicts)


def _gis_parcels_config(mapping: dict[str, Any] | None) -> tuple[str, str, str, str]:
    """Layer + column names from mapping.gisParcels (defaults match Arapahoe Open GIS)."""
    cfg = (mapping or {}).get("gisParcels") or {}
    layer = _strip(cfg.get("layer")) or _DEFAULT_GIS_LAYER
    pin_col = _strip(cfg.get("pin")) or _DEFAULT_GIS_PIN_COL
    code_col = _strip(cfg.get("neighborhoodCode")) or _DEFAULT_GIS_CODE_COL
    name_col = _strip(cfg.get("neighborhood")) or _DEFAULT_GIS_NAME_COL
    return layer, pin_col, code_col, name_col


def read_neighborhood_by_pin(
    gdb_path: Path,
    *,
    mapping: dict[str, Any] | None = None,
    pin_digits: int = 9,
) -> dict[str, dict[str, str]]:
    try:
        import pyogrio
    except ImportError as exc:
        raise RuntimeError(
            f"Open GIS Parcels GDB present ({gdb_path}) but pyogrio is not installed. "
            "Run: pip install -r tools/requirements.txt"
        ) from exc

    layer, pin_col, code_col, name_col = _gis_parcels_config(mapping)
    df = pyogrio.read_dataframe(
        gdb_path,
        layer=layer,
        columns=[pin_col, code_col, name_col],
        read_geometry=False,
    )
    rows: list[tuple[Any, Any, Any]] = list(
        zip(
            df[pin_col].tolist(),
            df[code_col].tolist(),
            df[name_col].tolist(),
        )
    )
    out, n_conflict = neighborhood_by_pin_from_rows(rows, pin_digits=pin_digits)
    if n_conflict:
        print(
            f"Open GIS Parcels: omitted {n_conflict} PIN(s) with conflicting "
            "neighborhood code/name rows",
            file=sys.stderr,
        )
    return out


def attach_neighborhood_from_gis(
    parcel_record_map: dict[str, dict[str, Any]],
    neighborhood_by_pin: dict[str, dict[str, str]],
) -> int:
    n = 0
    for pin, rec in parcel_record_map.items():
        nb = neighborhood_by_pin.get(pin)
        if not nb:
            continue
        rec["neighborhood"] = nb["neighborhood"]
        rec["neighborhoodCode"] = nb["neighborhoodCode"]
        n += 1
    return n


def read_gis_parcels_data_as_of(gdb_path: Path) -> str | None:
    path = gdb_path.parent / GIS_PARCELS_DATA_AS_OF_FILENAME
    if not path.is_file():
        return None
    text = path.read_text(encoding="utf-8", errors="replace")
    first = _strip(text.splitlines()[0] if text.splitlines() else "")
    if not first:
        return None
    try:
        datetime.strptime(first, "%Y-%m-%d")
    except ValueError:
        print(
            f"Ignoring {path}: expected one YYYY-MM-DD line, got {first!r}",
            file=sys.stderr,
        )
        return None
    return first


def attach_state_use_label(
    rec: dict[str, Any],
    state_class_by_code: dict[str, str],
) -> bool:
    code = normalize_state_use_cd(str(rec.get("stateUseCd") or ""))
    if not code:
        return False
    if code != _strip(str(rec.get("stateUseCd") or "")):
        rec["stateUseCd"] = code
    label = state_class_by_code.get(code)
    if not label:
        return False
    rec["stateUseLabel"] = label
    return True


def enrich_parcel_record_from_sibling_marts(
    parcel_record_map: dict[str, dict[str, Any]],
    mapping: dict[str, Any],
    *,
    legal_descriptions_path: Path | None,
    legal_parties_path: Path | None,
    land_path: Path | None = None,
    building_path: Path | None = None,
    building_xfob_path: Path | None = None,
    transfers_path: Path | None = None,
    permits_path: Path | None = None,
    state_class_xlsx_path: Path | None = None,
    nbhd_xlsx_path: Path | None = None,
    gis_parcels_gdb_path: Path | None = None,
    ownership_path: Path | None = None,
    subdivision_path: Path | None = None,
    values_path: Path | None = None,
    filing_path: Path | None = None,
    parcels_csv_path: Path | None = None,
) -> dict[str, int]:
    """Merge sibling mart tables / GIS neighborhood into parcel-record rows."""
    del building_xfob_path  # reserved; not used on PPINum layout today
    pin_digits = int(mapping.get("identifierDigits", 9))

    legal_by_pin = (
        read_legal_description_display_by_pin(
            legal_descriptions_path, mapping, pin_digits=pin_digits
        )
        if legal_descriptions_path and legal_descriptions_path.is_file()
        else {}
    )
    ownership_by_pin = (
        read_ownership_type_by_pin(legal_parties_path, mapping, pin_digits=pin_digits)
        if legal_parties_path and legal_parties_path.is_file()
        else {}
    )
    ownership_name_by_pin = (
        read_ownership_name_fields_by_pin(
            ownership_path, mapping, pin_digits=pin_digits
        )
        if ownership_path and ownership_path.is_file()
        else {}
    )
    subdivision_by_pin = (
        read_subdivision_fields_by_pin(
            subdivision_path, mapping, pin_digits=pin_digits
        )
        if subdivision_path and subdivision_path.is_file()
        else {}
    )
    land_by_pin = (
        read_land_fields_by_pin(land_path, mapping, pin_digits=pin_digits)
        if land_path
        else {}
    )
    building_by_pin = (
        read_building_fields_by_pin(building_path, mapping, pin_digits=pin_digits)
        if building_path
        else {}
    )
    transfers_by_pin = (
        read_transfers_by_pin(transfers_path, mapping, pin_digits=pin_digits)
        if transfers_path
        else {}
    )
    permits_by_pin = (
        read_permits_by_pin(permits_path, mapping, pin_digits=pin_digits)
        if permits_path
        else {}
    )
    values_by_pin = (
        read_values_parcel_enrichment_by_pin(
            values_path, mapping, pin_digits=pin_digits
        )
        if values_path and values_path.is_file()
        else {}
    )
    filing_lookup = (
        read_filing_lookup_by_recording_no(filing_path, mapping)
        if filing_path and filing_path.is_file()
        else {}
    )
    parcels_csv_by_pin = (
        read_parcels_csv_enrichment_by_pin(
            parcels_csv_path, mapping, pin_digits=pin_digits
        )
        if parcels_csv_path and parcels_csv_path.is_file()
        else {}
    )
    state_class_by_code = (
        read_state_class_description_by_code(state_class_xlsx_path)
        if state_class_xlsx_path
        else {}
    )
    neighborhood_by_pin: dict[str, dict[str, str]] = {}
    if gis_parcels_gdb_path and gis_parcels_gdb_path.exists():
        neighborhood_by_pin = read_neighborhood_by_pin(
            gis_parcels_gdb_path, mapping=mapping, pin_digits=pin_digits
        )
        print(
            f"Open GIS Parcels neighborhood map: {len(neighborhood_by_pin)} PINs",
            file=sys.stderr,
        )
    elif nbhd_xlsx_path and nbhd_xlsx_path.is_file():
        nbhd_count = len(read_nbhd_description_by_code(nbhd_xlsx_path))
        if nbhd_count:
            print(
                f"NBHD lookup loaded ({nbhd_count} codes); not joined - "
                "place Assessor_Parcels Open GIS GDB under supporting-data/county-gis/",
                file=sys.stderr,
            )

    counts = {
        "legalDescrDisplay": 0,
        "ownershipType": 0,
        "ownershipName": 0,
        "subdivision": 0,
        "land": 0,
        "building": 0,
        "transfers": 0,
        "permits": 0,
        "stateUseLabel": 0,
        "neighborhood": 0,
        "valuesDetail": 0,
        "filing": 0,
        "parcelsCsv": 0,
    }
    for pin, rec in parcel_record_map.items():
        mart_legal = legal_by_pin.get(pin)
        if mart_legal:
            rec["legalDescrDisplay"] = mart_legal
            counts["legalDescrDisplay"] += 1
        ownership = ownership_by_pin.get(pin)
        if ownership:
            rec["ownershipType"] = ownership
            counts["ownershipType"] += 1
        ownership_name = ownership_name_by_pin.get(pin)
        if ownership_name:
            rec.update(ownership_name)
            counts["ownershipName"] += 1
        subdivision = subdivision_by_pin.get(pin)
        if subdivision:
            if subdivision.get("subdivisionName") and not rec.get("subdivisionName"):
                rec["subdivisionName"] = subdivision["subdivisionName"]
            if subdivision.get("subdivisionCd") and not rec.get("subdivisionCd"):
                rec["subdivisionCd"] = subdivision["subdivisionCd"]
            for field in ("lotNo", "blockNo", "tractNo"):
                if subdivision.get(field) and not rec.get(field):
                    rec[field] = subdivision[field]
            recording = subdivision.get("subdivisionCd")
            if recording and filing_lookup.get(recording):
                filing = filing_lookup[recording]
                if filing.get("filingDescr") and not rec.get("filingDescr"):
                    rec["filingDescr"] = filing["filingDescr"]
                if filing.get("filingNo") and not rec.get("filingNo"):
                    rec["filingNo"] = filing["filingNo"]
                if filing.get("subdivisionName") and not rec.get("subdivisionName"):
                    rec["subdivisionName"] = filing["subdivisionName"]
                counts["filing"] += 1
            counts["subdivision"] += 1
        land = land_by_pin.get(pin)
        if land:
            rec.update(land)
            counts["land"] += 1
        building = building_by_pin.get(pin)
        if building:
            if building.get("landUse"):
                rec["landUse"] = building["landUse"]
            if building.get("buildings"):
                rec["buildings"] = building["buildings"]
            counts["building"] += 1
        transfers = transfers_by_pin.get(pin)
        if transfers:
            rec["transfers"] = transfers
            counts["transfers"] += 1
        permits = permits_by_pin.get(pin)
        if permits:
            rec["permits"] = permits
            counts["permits"] += 1
        if state_class_by_code and attach_state_use_label(rec, state_class_by_code):
            counts["stateUseLabel"] += 1
        values_detail = values_by_pin.get(pin)
        if values_detail:
            if values_detail.get("landLines") and not rec.get("landLines"):
                rec["landLines"] = values_detail["landLines"]
            if rec.get("landActual") is None and values_detail.get("landActual") is not None:
                rec["landActual"] = values_detail["landActual"]
            if (
                rec.get("improvementActual") is None
                and values_detail.get("improvementActual") is not None
            ):
                rec["improvementActual"] = values_detail["improvementActual"]
            if values_detail.get("stateUseCd") and not rec.get("stateUseCd"):
                rec["stateUseCd"] = values_detail["stateUseCd"]
            counts["valuesDetail"] += 1
        parcels_csv = parcels_csv_by_pin.get(pin)
        if parcels_csv:
            for field in ("blockNo", "tractNo", "filingDescr"):
                if parcels_csv.get(field) and not rec.get(field):
                    rec[field] = parcels_csv[field]
            if parcels_csv.get("legalDescrFull") and not rec.get("legalDescrFull"):
                full = parcels_csv["legalDescrFull"]
                rec["legalDescrFull"] = full
                rec["legalDescrDisplay"] = legal_descr_display_tail(full)
            counts["parcelsCsv"] += 1
    if neighborhood_by_pin:
        counts["neighborhood"] = attach_neighborhood_from_gis(
            parcel_record_map, neighborhood_by_pin
        )
    return counts


def print_parcel_record_shard_size_stats(shard_dir: Path) -> None:
    sizes = sorted(p.stat().st_size for p in shard_dir.glob("*.json"))
    if not sizes:
        return
    total_mb = sum(sizes) / (1024 * 1024)
    p90 = sizes[int(len(sizes) * 0.9)]
    p99 = sizes[int(len(sizes) * 0.99)]
    over_500 = sum(1 for s in sizes if s > 500 * 1024)
    over_1m = sum(1 for s in sizes if s > 1024 * 1024)
    print(
        f"Shard size stats: {len(sizes)} files, {total_mb:.2f} MiB total; "
        f"min {sizes[0] / 1024:.1f} KiB, median {statistics.median(sizes) / 1024:.1f} KiB, "
        f"mean {statistics.mean(sizes) / 1024:.1f} KiB, max {sizes[-1] / 1024:.1f} KiB; "
        f"p90 {p90 / 1024:.1f} KiB, p99 {p99 / 1024:.1f} KiB; "
        f">{500} KiB: {over_500}, >1 MiB: {over_1m}",
        file=sys.stderr,
    )


def write_parcel_record_shards(
    out_dir: Path,
    parcel_record_map: dict[str, dict[str, Any]],
    parcel_snapshot: dict[str, Any],
    *,
    pin_digits: int = 9,
    county_id: str = "arapahoe",
    separators: tuple[str, str] = (",", ":"),
) -> None:
    """Write plain JSON shards by account-id prefix under out_dir."""
    validate_out_dir(out_dir, ship=False)

    county = _strip(county_id) or "county"
    shard_dir = out_dir / f"{county}-parcel-record-by-pin"
    if shard_dir.exists():
        for old in shard_dir.glob("*.json.gz"):
            old.unlink()
        for old in shard_dir.glob("*.json"):
            old.unlink()
    shard_dir.mkdir(parents=True, exist_ok=True)

    legacy_mono = out_dir / f"{county}-parcel-record-by-pin.json.gz"
    if legacy_mono.exists():
        legacy_mono.unlink()
    legacy_mono_json = out_dir / f"{county}-parcel-record-by-pin.json"
    if legacy_mono_json.exists():
        legacy_mono_json.unlink()

    shards: dict[str, dict[str, dict[str, Any]]] = defaultdict(dict)
    for pin, row in parcel_record_map.items():
        if len(pin) < PARCEL_RECORD_SHARD_PREFIX_LEN:
            continue
        prefix = pin[:PARCEL_RECORD_SHARD_PREFIX_LEN]
        if not re.fullmatch(r"[A-Za-z0-9]+", prefix):
            continue
        shards[prefix][pin] = row

    for prefix in sorted(shards):
        by_pin = shards[prefix]
        path = shard_dir / f"{prefix}.json"
        path.write_text(
            json.dumps(
                {
                    "snapshot": parcel_snapshot,
                    "pinDigits": pin_digits,
                    "shardPrefix": prefix,
                    "byPin": by_pin,
                },
                separators=separators,
            ),
            encoding="utf-8",
        )

    total_bytes = sum((shard_dir / f"{prefix}.json").stat().st_size for prefix in shards)
    total_mb = total_bytes / (1024 * 1024)
    print(
        f"Wrote {shard_dir}/ ({len(shards)} shards, {len(parcel_record_map)} pins, "
        f"{total_mb:.2f} MiB total)",
        file=sys.stderr,
    )
    print_parcel_record_shard_size_stats(shard_dir)
