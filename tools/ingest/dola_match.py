#!/usr/bin/env python3
# Metro Tax Lookup - Arapahoe County
# Copyright (C) 2026 Jesse Lind
# SPDX-License-Identifier: AGPL-3.0-or-later
# See LICENSE for full terms or https://www.gnu.org/licenses/agpl-3.0.html

"""DOLA mill join for the new ingest engine.

Reimplements the old rebuild's DOLA Property Tax Entities join inside
tools/ingest/ only. Does not import from build_arapahoe_parcel_levy_index.py.

Mill source: supporting-data/dola/property-tax-entities-export.csv (tracked;
statewide Property Tax Entities export; ingest filters by certifying county).
Authority overrides (shared curated input, both engines): tools/arapahoe_dola_authority_overrides.json.
Does not import from build_arapahoe_parcel_levy_index.py.
"""

from __future__ import annotations

import csv
import json
import math
import re
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from rapidfuzz import fuzz

_INGEST_DIR = Path(__file__).resolve().parent
_TOOLS_DIR = _INGEST_DIR.parent
_REPO_ROOT = _TOOLS_DIR.parent
DEFAULT_DOLA_CSV = _REPO_ROOT / "supporting-data" / "dola" / "property-tax-entities-export.csv"
DEFAULT_DOLA_XLSX = _REPO_ROOT / "supporting-data" / "dola" / "property-tax-entities-export.xlsx"
# Shared with the shipping rebuild: curated mart↔DOLA reconciliation, not engine output.
DEFAULT_OVERRIDES = _TOOLS_DIR / "arapahoe_dola_authority_overrides.json"


def default_dola_export_path() -> Path:
    if DEFAULT_DOLA_CSV.is_file():
        return DEFAULT_DOLA_CSV
    return DEFAULT_DOLA_XLSX


def strip_field(s: str | None) -> str:
    if s is None:
        return ""
    return str(s).strip()


def parse_levy_mills_cell(val: Any) -> float | None:
    """Parse DOLA LGIS total levy cell; returns None if missing or invalid."""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        if isinstance(val, float) and math.isnan(val):
            return None
        return float(val)
    s = strip_field(str(val))
    if not s:
        return None
    try:
        return float(s.replace(",", ""))
    except ValueError:
        return None


def mart_line_looks_bond_purpose(authority_upper: str) -> bool:
    au = strip_field(authority_upper).upper()
    return "BOND" in au or "DEBT" in au


def dola_name_looks_bond_purpose(legal_name: str) -> bool:
    u = strip_field(legal_name).upper()
    if "BOND" in u:
        return True
    if "DEBT SERVICE" in u:
        return True
    return False


def normalize_for_match(name: str) -> str:
    """Normalize mart labels and DOLA legal names before fuzzy compare."""
    s = strip_field(name).upper()
    s = s.replace("&", " AND ")
    s = s.replace("#", " ")
    s = re.sub(r"[^A-Z0-9\s]", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    s = re.sub(r"\bDIST\b", "DISTRICT", s)
    s = re.sub(r"\bSCH\b", "SCHOOL", s)
    s = re.sub(r"\bVLG\b", "VILLAGE", s)
    s = re.sub(r"\bMD\b", "METROPOLITAN DISTRICT", s)
    s = re.sub(r"\bS SUBURBAN\b", "SOUTH SUBURBAN", s)
    s = s.replace("DISTRRICT", "DISTRICT")
    s = re.sub(r"\bMETRO\b", "METROPOLITAN", s)
    return s


def load_overrides(path: Path) -> dict[str, dict[str, Any]]:
    """Load authority overrides keyed by uppercase mart authority label."""
    if not path.is_file():
        return {}
    data = json.loads(path.read_text(encoding="utf-8"))
    raw = data.get("byAuthorityUpper") or {}
    return {strip_field(k).upper(): v for k, v in raw.items()}


def _te_id_str(x: Any) -> str:
    if x is None:
        return ""
    return strip_field(str(x))


def _coalesce_lg_id_from_entity(lg_id: Any, tax_entity_id: Any) -> Any:
    lid = strip_field(str(lg_id)) if lg_id is not None else ""
    if lid:
        return lid
    te = _te_id_str(tax_entity_id)
    if "/" in te:
        left = te.split("/", 1)[0].strip()
        if left.isdigit():
            return left
    return None


def build_entities_by_te_id(entities: list[dict[str, Any]]) -> dict[str, dict[str, Any]]:
    out: dict[str, dict[str, Any]] = {}
    for e in entities:
        tid = _te_id_str(e.get("taxEntityId"))
        if not tid:
            continue
        if tid in out:
            print(
                f"Duplicate Tax Entity ID in DOLA rows: {tid} (keeping first).",
                file=sys.stderr,
            )
            continue
        out[tid] = e
    return out


def attach_levy_mills(
    dola: dict[str, Any],
    authority_upper: str,
    entities_by_te_id: dict[str, dict[str, Any]],
) -> dict[str, Any]:
    """
    Attach mills from the matched DOLA row when bond/debt purpose aligns.
    Wrong mills are worse than missing: mismatch clears mills.
    """
    method = dola.get("method")
    if method in ("none", "skipped"):
        return dola
    te_str = _te_id_str(dola.get("taxEntityId"))
    if not te_str:
        out = dict(dola)
        out["mills"] = None
        return out
    ent = entities_by_te_id.get(te_str)
    matched_name = strip_field(str(dola.get("matchedLegalName") or ""))
    if not ent:
        out = dict(dola)
        out["mills"] = None
        return out
    legal = strip_field(str(ent.get("legalName") or ""))

    mart_bond = mart_line_looks_bond_purpose(authority_upper)
    dola_bond = dola_name_looks_bond_purpose(matched_name or legal)

    if mart_bond != dola_bond:
        out = dict(dola)
        out["mills"] = None
        out["millsReason"] = "bond_purpose_mismatch"
        return out

    levy = ent.get("levyMills")
    if levy is None:
        out = dict(dola)
        out["mills"] = None
        return out

    out = dict(dola)
    out["mills"] = round(float(levy), 6)
    return out


def enrich_overrides_from_entities(
    overrides: dict[str, dict[str, Any]],
    entities: list[dict[str, Any]],
    min_score: float = 0.88,
) -> dict[str, dict[str, Any]]:
    """Fill taxEntityId / lgId on override rows by matching legalName to DOLA rows."""
    out: dict[str, dict[str, Any]] = {}
    resolved = 0
    for key, v in overrides.items():
        ovr = dict(v)
        if ovr.get("taxEntityId"):
            out[key] = ovr
            continue
        ln = ovr.get("legalName")
        if not ln or not entities:
            out[key] = ovr
            continue
        q = normalize_for_match(str(ln))
        best: dict[str, Any] | None = None
        best_score = -1.0
        for e in entities:
            s = fuzz.token_sort_ratio(q, e["norm"]) / 100.0
            if s > best_score:
                best_score = s
                best = e
        if best and best_score >= min_score:
            ovr["taxEntityId"] = best.get("taxEntityId")
            ovr["lgId"] = best.get("lgId")
            ovr["resolvedFromXlsx"] = True
            ovr["resolvedScore"] = round(best_score, 4)
            resolved += 1
        out[key] = ovr
    if resolved:
        print(
            f"Enriched {resolved} override rows from DOLA export "
            f"(>= {min_score:.0%} name match).",
            file=sys.stderr,
        )
    return out


def _dola_column_indices(
    headers: list[str],
) -> tuple[int | None, int | None, int | None, int | None, int | None] | None:
    idx_name = None
    idx_entity = None
    idx_lgid = None
    idx_county = None
    idx_levy = None
    for i, h in enumerate(headers):
        hl = h.lower()
        if idx_name is None and "name" in hl and "tax" in hl:
            idx_name = i
        if idx_name is None and hl in ("tax entity name", "entity name", "legal name"):
            idx_name = i
        if idx_entity is None and "tax entity" in hl and "id" in hl:
            idx_entity = i
        if idx_lgid is None and ("lgid" in hl.replace(" ", "") or hl == "lg id"):
            idx_lgid = i
        if idx_county is None and "certifying" in hl and "county" in hl:
            idx_county = i
        if idx_levy is None and "levy" in hl and ("total" in hl or "budget" in hl):
            idx_levy = i

    if idx_name is None:
        for i, h in enumerate(headers):
            if "name" in h.lower() and "county" not in h.lower():
                idx_name = i
                break

    if idx_name is None:
        return None
    return idx_name, idx_entity, idx_lgid, idx_county, idx_levy


def _entities_from_dola_table_rows(
    rows: Any,
    idx_name: int,
    idx_entity: int | None,
    idx_lgid: int | None,
    idx_county: int | None,
    idx_levy: int | None,
    certifying_county_upper: str,
) -> list[dict[str, Any]]:
    entities: list[dict[str, Any]] = []
    for row in rows:
        if row is None:
            continue
        cells = list(row)
        if not cells or idx_name >= len(cells):
            continue
        if idx_county is not None and idx_county < len(cells):
            cty = strip_field(str(cells[idx_county] if cells[idx_county] is not None else ""))
            if cty.upper() != certifying_county_upper:
                continue
        legal = cells[idx_name]
        if legal is None or strip_field(str(legal)) == "":
            continue
        legal_s = strip_field(str(legal))
        te_id = ""
        if idx_entity is not None and idx_entity < len(cells) and cells[idx_entity] is not None:
            te_id = strip_field(str(cells[idx_entity]))
        lg = ""
        if idx_lgid is not None and idx_lgid < len(cells) and cells[idx_lgid] is not None:
            lg = strip_field(str(cells[idx_lgid]))
        levy_mills: float | None = None
        if idx_levy is not None and idx_levy < len(cells):
            levy_mills = parse_levy_mills_cell(cells[idx_levy])
        norm = normalize_for_match(legal_s)
        if not norm:
            continue
        entities.append(
            {
                "legalName": legal_s,
                "norm": norm,
                "taxEntityId": te_id or None,
                "lgId": lg or None,
                "levyMills": levy_mills,
            }
        )
    return entities


def load_dola_entities_csv(
    csv_path: Path, certifying_county: str = "Arapahoe"
) -> tuple[list[dict[str, Any]], str | None, bool]:
    ccu = strip_field(certifying_county).upper()
    if not ccu:
        print("DOLA CSV: empty certifying county; skipping DOLA join.", file=sys.stderr)
        return [], None, False
    with csv_path.open(newline="", encoding="utf-8-sig", errors="replace") as f:
        reader = csv.reader(f)
        try:
            header_row = next(reader)
        except StopIteration:
            return [], None, False
        headers = [strip_field(str(h) if h is not None else "") for h in header_row]
        parsed = _dola_column_indices(headers)
        if parsed is None:
            print("Could not detect name column in DOLA CSV; skipping DOLA join.", file=sys.stderr)
            return [], None, False
        idx_name, idx_entity, idx_lgid, idx_county, idx_levy = parsed
        county_filter_applied = idx_county is not None
        if idx_county is None:
            print(
                "DOLA CSV: no Certifying County column; using all rows "
                "(may duplicate Tax Entity IDs across counties).",
                file=sys.stderr,
            )
        levy_header = headers[idx_levy] if idx_levy is not None else None
        entities = _entities_from_dola_table_rows(
            reader, idx_name, idx_entity, idx_lgid, idx_county, idx_levy, ccu
        )
        return entities, levy_header, county_filter_applied


def load_dola_entities_xlsx(
    xlsx_path: Path, certifying_county: str = "Arapahoe"
) -> tuple[list[dict[str, Any]], str | None, bool]:
    try:
        from openpyxl import load_workbook
    except ImportError:
        print("openpyxl required for DOLA xlsx; skipping DOLA join.", file=sys.stderr)
        return [], None, False

    ccu = strip_field(certifying_county).upper()
    if not ccu:
        print("DOLA xlsx: empty certifying county; skipping DOLA join.", file=sys.stderr)
        return [], None, False
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    try:
        ws = wb.active
        rows = ws.iter_rows(values_only=True)
        try:
            header_row = next(rows)
        except StopIteration:
            return [], None, False
        headers = [strip_field(str(h) if h is not None else "") for h in header_row]
        parsed = _dola_column_indices(headers)
        if parsed is None:
            print("Could not detect name column in DOLA xlsx; skipping DOLA join.", file=sys.stderr)
            return [], None, False
        idx_name, idx_entity, idx_lgid, idx_county, idx_levy = parsed
        county_filter_applied = idx_county is not None
        if idx_county is None:
            print(
                "DOLA xlsx: no Certifying County column; using all rows "
                "(may duplicate Tax Entity IDs across counties).",
                file=sys.stderr,
            )
        levy_header = headers[idx_levy] if idx_levy is not None else None
        entities = _entities_from_dola_table_rows(
            rows, idx_name, idx_entity, idx_lgid, idx_county, idx_levy, ccu
        )
        return entities, levy_header, county_filter_applied
    finally:
        wb.close()


def load_dola_entities(
    path: Path, certifying_county: str = "Arapahoe"
) -> tuple[list[dict[str, Any]], str | None, bool]:
    if not path.is_file():
        return [], None, False
    suf = path.suffix.lower()
    if suf == ".csv":
        return load_dola_entities_csv(path, certifying_county)
    if suf in (".xlsx", ".xlsm"):
        return load_dola_entities_xlsx(path, certifying_county)
    print(f"Unsupported DOLA export format (expected .csv or .xlsx): {path}", file=sys.stderr)
    return [], None, False


def match_dola_line(
    authority_upper: str,
    entities: list[dict[str, Any]],
    overrides: dict[str, dict[str, Any]],
) -> dict[str, Any]:
    """Resolve one mart authority line to a DOLA tax entity (override, fuzzy, or none)."""
    au = strip_field(authority_upper).upper()
    ovr = overrides.get(au)
    query_base = normalize_for_match(authority_upper)
    if ovr and ovr.get("legalName"):
        query_base = normalize_for_match(str(ovr["legalName"]))

    def ura_extra() -> dict[str, Any]:
        if ovr and ovr.get("ura"):
            return {"uraHint": True}
        return {}

    if ovr and ovr.get("taxEntityId") and entities:
        want = _te_id_str(ovr.get("taxEntityId"))
        if want:
            for e in entities:
                if _te_id_str(e.get("taxEntityId")) == want:
                    return {
                        "method": "override",
                        "confidence": "high",
                        "taxEntityId": e.get("taxEntityId"),
                        "lgId": _coalesce_lg_id_from_entity(
                            e.get("lgId"), e.get("taxEntityId")
                        ),
                        "matchedLegalName": e.get("legalName"),
                        "score": 1.0,
                        **ura_extra(),
                    }

    if not entities:
        if ovr:
            return {
                "method": "override",
                "confidence": "low",
                "taxEntityId": ovr.get("taxEntityId"),
                "lgId": _coalesce_lg_id_from_entity(
                    ovr.get("lgId"), ovr.get("taxEntityId")
                ),
                "matchedLegalName": ovr.get("legalName"),
                "score": None,
                **ura_extra(),
            }
        return {"method": "none", "confidence": "low", "matchedLegalName": None, "score": None}

    best: dict[str, Any] | None = None
    best_score = -1.0
    for e in entities:
        score = fuzz.token_sort_ratio(query_base, e["norm"]) / 100.0
        if score > best_score:
            best_score = score
            best = e

    assert best is not None
    if best_score >= 0.92:
        conf = "high"
    elif best_score >= 0.78:
        conf = "medium"
    else:
        conf = "low"

    if ovr and ovr.get("legalName") and best_score < 0.78:
        query2 = normalize_for_match(str(ovr["legalName"]))
        for e in entities:
            score = fuzz.token_sort_ratio(query2, e["norm"]) / 100.0
            if score > best_score:
                best_score = score
                best = e
        if best_score >= 0.92:
            conf = "high"
        elif best_score >= 0.78:
            conf = "medium"
        else:
            conf = "low"

    if best_score < 0.70:
        return {
            "method": "none",
            "confidence": "low",
            "taxEntityId": None,
            "lgId": None,
            "matchedLegalName": None,
            "score": round(best_score, 4),
            **ura_extra(),
        }

    method = "fuzzy"
    if ovr:
        method = "override"

    return {
        "method": method,
        "confidence": conf,
        "taxEntityId": best.get("taxEntityId"),
        "lgId": _coalesce_lg_id_from_entity(
            best.get("lgId"), best.get("taxEntityId")
        ),
        "matchedLegalName": best.get("legalName"),
        "score": round(best_score, 4),
        **ura_extra(),
    }


def dola_match_for_mart_line(
    line_code: str,
    authority_upper: str,
    entities: list[dict[str, Any]],
    overrides: dict[str, dict[str, Any]],
    entities_by_te_id: dict[str, dict[str, Any]],
) -> dict[str, Any]:
    """ASSRFEES is an assessor fee row, not a taxing district."""
    code = strip_field(line_code).upper()
    if code == "ASSRFEES":
        return {
            "method": "skipped",
            "confidence": "high",
            "skipReason": "assessor_fee",
            "taxEntityId": None,
            "lgId": None,
            "matchedLegalName": None,
            "score": None,
        }
    au = strip_field(authority_upper).upper()
    ovr = overrides.get(au)
    dola = match_dola_line(authority_upper, entities, overrides)
    result = attach_levy_mills(dola, authority_upper, entities_by_te_id)
    if ovr is not None and ovr.get("millsOverride") is not None:
        try:
            mv = float(ovr["millsOverride"])
        except (TypeError, ValueError):
            return result
        out = dict(result)
        prev = out.get("mills")
        if isinstance(prev, (int, float)):
            out["dolaMills"] = round(float(prev), 6)
        out["mills"] = round(mv, 6)
        out["millsReason"] = "county_levy_table_override"
        return out
    return result


@dataclass
class DolaJoinContext:
    """Loaded DOLA entities + overrides for one ingest run."""

    entities: list[dict[str, Any]]
    entities_by_te_id: dict[str, dict[str, Any]]
    overrides: dict[str, dict[str, Any]]
    dola_path: Path | None
    levy_column_header: str | None
    certifying_county: str | None
    county_filter_applied: bool

    def match_line(self, line_code: str, authority_name: str) -> dict[str, Any]:
        return dola_match_for_mart_line(
            line_code,
            strip_field(authority_name).upper(),
            self.entities,
            self.overrides,
            self.entities_by_te_id,
        )

    def snapshot_fields(self) -> dict[str, Any]:
        return {
            "dolaSource": self.dola_path.name if self.dola_path and self.dola_path.is_file() else None,
            "dolaRowCount": len(self.entities),
            "dolaCertifyingCounty": (
                self.certifying_county if self.county_filter_applied else None
            ),
            "dolaLevyColumn": self.levy_column_header,
        }


def load_dola_join_context(
    *,
    dola_export: Path | None = None,
    overrides_path: Path | None = None,
    certifying_county: str = "Arapahoe",
) -> DolaJoinContext:
    """Load DOLA CSV/xlsx + new-engine overrides for mill join."""
    dola_path = dola_export if dola_export is not None else default_dola_export_path()
    ovr_path = overrides_path if overrides_path is not None else DEFAULT_OVERRIDES
    cc = strip_field(certifying_county) or "Arapahoe"

    entities, levy_col, county_filter = load_dola_entities(dola_path, cc)
    overrides = load_overrides(ovr_path)
    if entities:
        overrides = enrich_overrides_from_entities(overrides, entities)
    by_te = build_entities_by_te_id(entities)

    if entities:
        if county_filter:
            print(
                f"DOLA entities loaded: {len(entities)} (certifying county {cc} only)",
                file=sys.stderr,
            )
        else:
            print(f"DOLA entities loaded: {len(entities)}", file=sys.stderr)
    elif dola_path.is_file():
        print(f"DOLA export present but no entities loaded: {dola_path}", file=sys.stderr)
    else:
        print(
            f"DOLA export not found ({dola_path}); levy lines get method=none.",
            file=sys.stderr,
        )

    return DolaJoinContext(
        entities=entities,
        entities_by_te_id=by_te,
        overrides=overrides,
        dola_path=dola_path if dola_path.is_file() else None,
        levy_column_header=levy_col,
        certifying_county=cc,
        county_filter_applied=county_filter,
    )
