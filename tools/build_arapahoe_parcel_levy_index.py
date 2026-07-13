#!/usr/bin/env python3
# Metro Tax Lookup - Arapahoe County
# Copyright (C) 2026 Jesse Lind
# SPDX-License-Identifier: AGPL-3.0-or-later
# See LICENSE for full terms or https://www.gnu.org/licenses/agpl-3.0.html

"""
Build Arapahoe parcel → levy stack index from county datamart CSV exports.

Outputs (default: metro-tax-lookup/public/data/):
  - arapahoe-levy-stacks-by-tag-id.json — TAGId → levy lines (+ DOLA match, mills from LGIS export when safe)
  - arapahoe-pin-to-tag.json — Pin → { tagId, tagShortDescr, ain, … } (large; see --skip-pin-map)
  - arapahoe-situs-to-pins.json — situs lookup key → [{ pin, label }, ...] for home address flow (see --skip-pin-map)
  - arapahoe-parcel-record-by-pin/<prefix>.json — county-record fields sharded by PIN prefix
    (see PARCEL_RECORD_SHARD_PREFIX_LEN; Main Parcel + sibling mart joins; lazy load after levy;
    see --skip-pin-map)

Parcel-record sibling inputs (optional; skipped if missing):
  Mart_DescrHeader, Mart_LegalParty, Mart_RDE_LndAll, Mart_RDE_BLD, Mart_RDE_Xfob,
  Mart_Transfers (Sale rows with Book+Page), Mart_RDE_Permit,
  State Class Codes xlsx → stateUseLabel, NBHD codes xlsx (loaded only; not joined until Main Parcel
  exposes a neighborhood code).

Mart_TA_TAG: supporting-data/county-mart/.../Tax Authority Groups and Tax Authorities.csv
Main parcel: supporting-data/county-mart/.../Main Parcel Table.csv
Optional DOLA: supporting-data/dola/property-tax-entities-export.csv or .xlsx (LGIS Property Tax Entities:
  https://dola.colorado.gov/dlg_lgis_ui_pu/publicLGTaxEntities.jsf — canonical key src/lib/dataSourceUrls.ts DOLA_LGIS_PROPERTY_TAX_ENTITIES)

Run from repo root:
  cd metro-tax-lookup && source .venv/bin/activate && pip install -r tools/requirements.txt
  python tools/build_arapahoe_parcel_levy_index.py

Maintainer notes:
  - Levy.aspx?id= uses TAGId (tax area), same as Mart_TA_TAG Field2 and Main Parcel TAGId.
    It is not a per-parcel serial; many parcels share one TAGId.
  - Field5 code ASSRFEES is the county assessor fee in the mart export; it is not shown on the
    county online Tax District Levies page. The app PIN-load path omits
    that row so the list matches the table users copy from.
  - UI Book Page links use Clerk & Recorder public search (Book+Page concatenated); build stores
    display bookPage only (no URL in JSON).
"""

from __future__ import annotations

import argparse
import csv
import json
import math
import re
import statistics
import sys
from collections import defaultdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

try:
    from rapidfuzz import fuzz
except ImportError:
    print("Install dependencies: pip install -r tools/requirements.txt", file=sys.stderr)
    raise

REPO_ROOT = Path(__file__).resolve().parent.parent
SUPPORTING_DATA = REPO_ROOT / "supporting-data"
COUNTY_MART = SUPPORTING_DATA / "county-mart"
# One line YYYY-MM-DD: date you downloaded / refreshed this mart batch (not auto-detected).
COUNTY_DATA_AS_OF_FILE = "data-as-of.txt"
DOLA_DIR = SUPPORTING_DATA / "dola"

DEFAULT_MAIN = (
    COUNTY_MART
    / "Main Parcel Table (CSV)"
    / "Main Parcel Table.csv"
)
DEFAULT_MART = (
    COUNTY_MART
    / "Tax Authority Groups and Tax Authorities (CSV)"
    / "Tax Authority Groups and Tax Authorities.csv"
)
DEFAULT_LEGAL_DESCRIPTIONS = (
    COUNTY_MART
    / "Parcel Legal Descriptions (CSV)"
    / "Parcel Legal Descriptions.csv"
)
DEFAULT_LEGAL_PARTIES = (
    COUNTY_MART
    / "Parcel Legal Parties (CSV)"
    / "Parcel Legal Parties.csv"
)
DEFAULT_LAND = (
    COUNTY_MART
    / "Parcel Land Information (CSV)"
    / "Parcel Land Information.csv"
)
DEFAULT_BUILDING = (
    COUNTY_MART
    / "Parcel Building Information (CSV)"
    / "Parcel Building Information.csv"
)
DEFAULT_BUILDING_XFOB = (
    COUNTY_MART
    / "Parcel Building Extra Features (CSV)"
    / "Parcel Building Extra Features.csv"
)
DEFAULT_TRANSFERS = (
    COUNTY_MART
    / "Parcel Transfer Information (CSV)"
    / "Parcel Transfer Information.csv"
)
DEFAULT_PERMITS = (
    COUNTY_MART
    / "Parcel Permit Information (CSV)"
    / "Parcel Permit Information.csv"
)
DEFAULT_NBHD_XLSX = (
    COUNTY_MART / "Main Parcel Table (CSV)" / "NBHD codes 4 2017.xlsx"
)
DEFAULT_STATE_CLASS_XLSX = (
    COUNTY_MART / "Main Parcel Table (CSV)" / "State Class Codes 3 30 2015.xlsx"
)
DEFAULT_DOLA_CSV = DOLA_DIR / "property-tax-entities-export.csv"
DEFAULT_DOLA_XLSX = DOLA_DIR / "property-tax-entities-export.xlsx"


def default_dola_export_path() -> Path:
    """Prefer committed CSV; fall back to local xlsx when CSV is absent."""
    if DEFAULT_DOLA_CSV.is_file():
        return DEFAULT_DOLA_CSV
    return DEFAULT_DOLA_XLSX
DEFAULT_OVERRIDES = Path(__file__).resolve().parent / "arapahoe_dola_authority_overrides.json"
DEFAULT_OUT_DIR = REPO_ROOT / "public" / "data"

LEVY_ASPX_BASE = "https://parcelsearch.arapahoegov.com/Levy.aspx?id="


def strip_field(s: str | None) -> str:
    """Return a trimmed string, treating None as empty."""
    if s is None:
        return ""
    return str(s).strip()


def parse_parcel_value_cell(val: Any) -> float | None:
    """Parse TotalActual / TotalAssessed from Main Parcel CSV; returns None if missing or invalid."""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        if isinstance(val, float) and math.isnan(val):
            return None
        return float(val)
    s = strip_field(str(val))
    if not s:
        return None
    try:
        return float(s.replace(",", ""))
    except ValueError:
        return None


# Colorado DPT assessed rates for residential property (2025+). School columns match
# round(actual x school rate) per component; local land uses round(land actual x local rate)
# and local building is totalAssessed minus local land (county PPINum.aspx pattern).
COLORADO_SCHOOL_ASSESSED_RATE = 0.0705
COLORADO_LOCAL_ASSESSED_RATE = 0.068
DUAL_ASSESSED_MIN_ASSESSMENT_YEAR = 2025


def parse_assessment_year_cell(val: Any) -> int | None:
    """Parse Main Parcel AssessmentYear; returns None if missing or invalid."""
    s = strip_field(str(val)) if val is not None else ""
    if not s:
        return None
    try:
        year = int(float(s))
    except ValueError:
        return None
    return year if 1900 <= year <= 2100 else None


def parcel_row_qualifies_for_dual_assessed_splits(row: dict[str, str]) -> bool:
    """Local assessed building/land splits apply to real property from 2025."""
    year = parse_assessment_year_cell(row.get("AssessmentYear"))
    if year is None or year < DUAL_ASSESSED_MIN_ASSESSMENT_YEAR:
        return False
    return strip_field(row.get("TaxRollDescr", "")).upper() == "REAL"


def parcel_row_qualifies_for_school_assessed_splits(row: dict[str, str]) -> bool:
    """School assessed splits apply to improved residential (county Improvement class) from 2025."""
    if not parcel_row_qualifies_for_dual_assessed_splits(row):
        return False
    return strip_field(row.get("PropertyClassDescr", "")) == "Improvement"


def _positive_actual(val: float | None) -> float:
    if val is None or not math.isfinite(val) or val <= 0:
        return 0.0
    return float(val)


def round_school_assessed_component(actual: float | None) -> int | None:
    if actual is None or not math.isfinite(actual):
        return None
    return round(actual * COLORADO_SCHOOL_ASSESSED_RATE)


def school_assessed_fields_from_actuals(
    improvement_actual: float | None,
    land_actual: float | None,
    total_actual: float | None,
) -> dict[str, int]:
    """County-style school assessed: round each actual component x school rate, then sum."""
    building = round_school_assessed_component(improvement_actual)
    land = round_school_assessed_component(land_actual)
    if building is not None or land is not None:
        out: dict[str, int] = {}
        if building is not None:
            out["schoolAssessedBuilding"] = building
        if land is not None:
            out["schoolAssessedLand"] = land
        out["schoolAssessedTotal"] = (building or 0) + (land or 0)
        return out
    total = round_school_assessed_component(total_actual)
    if total is not None:
        return {"schoolAssessedTotal": total}
    return {}


def local_assessed_split_fields(
    improvement_actual: float | None,
    land_actual: float | None,
    total_assessed: float | None,
) -> dict[str, int]:
    """County-style local assessed building/land (totalAssessed comes from mart)."""
    if total_assessed is None or not math.isfinite(total_assessed):
        return {}
    total_int = round(total_assessed)
    imp = _positive_actual(improvement_actual)
    land = _positive_actual(land_actual)
    if imp == 0 and land > 0:
        return {"assessedBuilding": 0, "assessedLand": total_int}
    if imp > 0 and land == 0:
        return {"assessedBuilding": total_int, "assessedLand": 0}
    if imp > 0 and land > 0:
        land_assessed = round(land * COLORADO_LOCAL_ASSESSED_RATE)
        return {
            "assessedLand": land_assessed,
            "assessedBuilding": max(0, total_int - land_assessed),
        }
    return {}


def attach_computed_assessed_values(rec: dict[str, Any], row: dict[str, str]) -> None:
    """Add school and local assessed splits when DPT dual-rate rules apply."""
    if parcel_row_qualifies_for_dual_assessed_splits(row):
        rec.update(
            local_assessed_split_fields(
                rec.get("improvementActual"),
                rec.get("landActual"),
                rec.get("totalAssessed"),
            )
        )
    if parcel_row_qualifies_for_school_assessed_splits(row):
        rec.update(
            school_assessed_fields_from_actuals(
                rec.get("improvementActual"),
                rec.get("landActual"),
                rec.get("totalActual"),
            )
        )


def parse_levy_mills_cell(val: Any) -> float | None:
    """Parse DOLA LGIS total levy cell; returns None if missing or invalid."""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        if isinstance(val, float) and math.isnan(val):
            return None
        return float(val)
    s = strip_field(str(val))
    if not s:
        return None
    try:
        return float(s.replace(",", ""))
    except ValueError:
        return None


def mart_line_looks_bond_purpose(authority_upper: str) -> bool:
    """True if the mart label names a bond/debt levy line (not general 'debt' in unrelated words)."""
    au = strip_field(authority_upper).upper()
    if "BOND" in au:
        return True
    if "DEBT" in au:
        return True
    return False


def dola_name_looks_bond_purpose(legal_name: str) -> bool:
    """True when a DOLA legal name reads like bond or debt service."""
    u = strip_field(legal_name).upper()
    if "BOND" in u:
        return True
    if "DEBT SERVICE" in u:
        return True
    return False


def normalize_pin(raw: str) -> str:
    """Digits-only PIN, exactly nine digits (zero-padded) for map keys."""
    digits = re.sub(r"\D", "", raw)
    if not digits:
        return ""
    return digits.zfill(9)[:9]


# Situs lookup keys must stay in sync with src/lib/arapahoeSitusLookup.ts (home address flow).
_STREET_DIR_TOKENS = frozenset(
    {
        "N",
        "S",
        "E",
        "W",
        "NE",
        "NW",
        "SE",
        "SW",
        "NORTH",
        "SOUTH",
        "EAST",
        "WEST",
        "NORTHEAST",
        "NORTHWEST",
        "SOUTHEAST",
        "SOUTHWEST",
    }
)
_STREET_TYPE_TOKENS = frozenset(
    {
        "ST",
        "STREET",
        "AVE",
        "AVENUE",
        "RD",
        "ROAD",
        "BLVD",
        "BOULEVARD",
        "DR",
        "DRIVE",
        "LN",
        "LANE",
        "CT",
        "COURT",
        "CIR",
        "CIRCLE",
        "WAY",
        "PL",
        "PLACE",
        "PKWY",
        "PARKWAY",
        "TRL",
        "TRAIL",
        "LOOP",
        "TER",
        "TERR",
        "TERRACE",
        "TPKE",
        "TURNPIKE",
        "HWY",
        "HIGHWAY",
        "BL",
        "PATH",
        "PLZ",
        "PLAZA",
        "RUN",
        "COVE",
        "PASS",
        "ALLEY",
        "ALY",
        "BEND",
        "XING",
        "CROSSING",
        "POINT",
        "PT",
        "COMMONS",
        "MALL",
    }
)


def normalize_street_name_key(raw: str) -> str:
    """County address search omits directionals and street types; mirror that on mart situs."""
    s = strip_field(raw).upper()
    if not s:
        return ""
    tokens = [t for t in re.split(r"[^\w]+", s) if t]
    kept: list[str] = []
    for t in tokens:
        if t in _STREET_DIR_TOKENS or t in _STREET_TYPE_TOKENS:
            continue
        kept.append(t)
    return " ".join(kept)


def normalize_street_number_key(primary: str, range_or_suffix: str) -> str:
    """Merge SAAddrNumber + optional SAStreetNumberSfx (e.g. 1/2); keep in sync with arapahoeSitusLookup.ts."""
    a = strip_field(primary)
    b = strip_field(range_or_suffix)
    merged = " ".join(x for x in (a, b) if x)
    if not merged:
        return ""
    merged_u = merged.upper().replace(" ", "")
    return "".join(c for c in merged_u if c.isdigit() or c in "/-")


def normalize_unit_key(raw: str) -> str:
    """Uppercase unit string with non-alphanumerics removed (situs lookup)."""
    s = strip_field(raw).upper()
    if not s:
        return ""
    return re.sub(r"[^A-Z0-9]", "", s)


def row_situs_lookup_key(row: dict[str, str]) -> str | None:
    """Stable num|name|unit key for one Main Parcel row, or None if unusable."""
    num = normalize_street_number_key(row.get("SAAddrNumber", ""), row.get("SAStreetNumberSfx", ""))
    name = normalize_street_name_key(row.get("SAStreetName", ""))
    unit = normalize_unit_key(row.get("SAUnitNumber", ""))
    if not num or not name:
        return None
    # Skip common placeholder situs rows in the mart (not useful for address search).
    if strip_field(row.get("SAAddrNumber", "")) == "0":
        return None
    if "TAG" in strip_field(row.get("SAStreetName", "")).upper():
        return None
    return f"{num}|{name}|{unit}"


def format_situs_label(row: dict[str, str]) -> str:
    """Human-readable situs line for UI labels (falls back to PIN)."""
    n = strip_field(row.get("SAAddrNumber", ""))
    pre = strip_field(row.get("SAPredirectional", ""))
    name = strip_field(row.get("SAStreetName", ""))
    typ = strip_field(row.get("SAStreetType", ""))
    post = strip_field(row.get("SAPostdirectional", ""))
    unit = strip_field(row.get("SAUnitNumber", ""))
    city = strip_field(row.get("SACity", ""))
    line1 = " ".join(x for x in (n, pre, name, typ, post) if x)
    if unit:
        line1 = f"{line1} Unit {unit}".strip()
    if city:
        return f"{line1}, {city}".strip()
    return line1 or strip_field(row.get("Pin", ""))


def _parcel_record_from_row(row: dict[str, str]) -> dict[str, Any]:
    legal_full = _optional_str(row, "LegalDescr")
    legal_display = legal_descr_display_tail(legal_full) if legal_full else None
    rec: dict[str, Any] = {
        "ain": _optional_str(row, "AIN"),
        "situsAddress": _optional_str(row, "SAFreeFormAddr"),
        "situsCity": _optional_str(row, "SACity"),
        "ownerList": _optional_str(row, "OwnerList"),
        "ownerDeliveryAddress": _optional_str(row, "CurDeliveryAddr"),
        "ownerCityStateZip": _optional_str(row, "CurLastLine"),
        "legalDescrFull": legal_full,
        "legalDescrDisplay": legal_display,
        "subdivisionCd": _optional_str(row, "SubdivisionCd"),
        "subdivisionName": _optional_str(row, "SubdivisionName"),
        "taxRollDescr": _optional_str(row, "TaxRollDescr"),
        "propertyClassDescr": _optional_str(row, "PropertyClassDescr"),
        "totalActual": parse_parcel_value_cell(row.get("TotalActual")),
        "improvementActual": parse_parcel_value_cell(row.get("ImprovementActual")),
        "landActual": parse_parcel_value_cell(row.get("LandActual")),
        "totalAssessed": parse_parcel_value_cell(row.get("TotalAssessed")),
        "stateUseCd": _optional_str(row, "StateUseCd"),
        "parcelTaxYear": _optional_str(row, "TaxYear"),
        "assessmentYear": _optional_str(row, "AssessmentYear"),
    }
    attach_computed_assessed_values(rec, row)
    return rec


def _pin_map_first_row(row: dict[str, str]) -> dict[str, Any] | None:
    tag_id = strip_field(row.get("TAGId", ""))
    if not tag_id:
        return None
    short_d = strip_field(row.get("TAGShortDescr", ""))
    rec: dict[str, Any] = {
        "tagId": tag_id,
        "tagShortDescr": short_d,
        "totalActual": parse_parcel_value_cell(row.get("TotalActual")),
        "totalAssessed": parse_parcel_value_cell(row.get("TotalAssessed")),
        "parcelTaxYear": strip_field(row.get("TaxYear", "")) or None,
        "assessmentYear": strip_field(row.get("AssessmentYear", "")) or None,
        "propertyClassDescr": strip_field(row.get("PropertyClassDescr", "")) or None,
    }
    owner_list = strip_field(row.get("OwnerList", ""))
    if owner_list:
        rec["ownerList"] = owner_list
    ain = strip_field(row.get("AIN", ""))
    if ain:
        rec["ain"] = ain
    return rec


def _accumulate_situs_row(
    by_key: dict[str, dict[str, str]],
    row: dict[str, str],
    pin: str,
) -> None:
    lk = row_situs_lookup_key(row)
    if not lk:
        return
    label = format_situs_label(row)
    if lk not in by_key:
        by_key[lk] = {}
    if pin not in by_key[lk]:
        by_key[lk][pin] = label


def read_main_parcel_maps(
    path: Path,
) -> tuple[dict[str, dict[str, Any]], dict[str, list[dict[str, str]]], dict[str, dict[str, Any]]]:
    """One pass over Main Parcel: pin map, situs index, and parcel record map."""
    pin_map: dict[str, dict[str, Any]] = {}
    situs_by_key: dict[str, dict[str, str]] = {}
    parcel_record_map: dict[str, dict[str, Any]] = {}
    with path.open(newline="", encoding="utf-8", errors="replace") as f:
        r = csv.DictReader(f)
        for row in r:
            pin = normalize_pin(strip_field(row.get("Pin", "")))
            if not pin:
                continue
            _accumulate_situs_row(situs_by_key, row, pin)
            if pin not in parcel_record_map:
                parcel_record_map[pin] = _parcel_record_from_row(row)
            if pin not in pin_map:
                first = _pin_map_first_row(row)
                if first:
                    pin_map[pin] = first
            else:
                ain = strip_field(row.get("AIN", ""))
                if ain:
                    if not pin_map[pin].get("ain"):
                        pin_map[pin]["ain"] = ain
                    if not parcel_record_map[pin].get("ain"):
                        parcel_record_map[pin]["ain"] = ain
    situs_out: dict[str, list[dict[str, str]]] = {}
    for k, pin_labels in situs_by_key.items():
        items = [{"pin": p, "label": pin_labels[p]} for p in sorted(pin_labels.keys())]
        situs_out[k] = items
    return pin_map, merge_aggregate_situs_keys(situs_out), parcel_record_map


def build_situs_to_pins(path: Path) -> dict[str, list[dict[str, str]]]:
    """Pin -> one label each; multiple parcels can share the same lookup key."""
    _, situs_map, _ = read_main_parcel_maps(path)
    return situs_map


def merge_aggregate_situs_keys(
    by_key: dict[str, list[dict[str, str]]],
) -> dict[str, list[dict[str, str]]]:
    """
    Add num|name| keys that union every pin from num|name|* so unit can stay optional
    (county treats Unit as optional).
    """
    parent_pins: dict[str, dict[str, str]] = {}
    for k, items in by_key.items():
        parts = k.split("|")
        if len(parts) != 3:
            continue
        num, name, _unit = parts
        if not num or not name:
            continue
        parent = f"{num}|{name}|"
        bucket = parent_pins.setdefault(parent, {})
        for it in items:
            pin = it.get("pin", "")
            if pin and pin not in bucket:
                bucket[pin] = it.get("label", "")
    merged = dict(by_key)
    for parent, pmap in parent_pins.items():
        agg_list = [{"pin": p, "label": pmap[p]} for p in sorted(pmap.keys())]
        if parent not in merged:
            merged[parent] = agg_list
        else:
            combined: dict[str, str] = {x["pin"]: x["label"] for x in merged[parent]}
            for it in agg_list:
                combined.setdefault(it["pin"], it["label"])
            merged[parent] = [{"pin": p, "label": combined[p]} for p in sorted(combined.keys())]
    return merged


def mart_row_maps(fieldnames: list[str] | None) -> dict[str, str]:
    """Map logical names to actual CSV header."""
    if not fieldnames:
        return {}
    fset = {strip_field(h) for h in fieldnames}
    if fset >= {"Field1", "Field2", "Field3", "Field4", "Field5", "Field6"}:
        return {
            "internal_id": "Field1",
            "tag_id": "Field2",
            "tax_year": "Field3",
            "kind": "Field4",
            "line_code": "Field5",
            "authority_name": "Field6",
            "effective_year": "Field7",
            "status": "Field8",
        }
    # Named export (future)
    lower = {strip_field(h).lower(): strip_field(h) for h in fieldnames}

    def pick(*candidates: str) -> str | None:
        """Return the first header present in fieldnames (case-insensitive)."""
        for c in candidates:
            if c.lower() in lower:
                return lower[c.lower()]
        return None

    m = {
        "tag_id": pick("TAGId", "TagId", "tag_id"),
        "tax_year": pick("TaxYear", "tax_year"),
        "line_code": pick("LevyLineCode", "line_code", "Code"),
        "authority_name": pick("AuthorityName", "authority_name", "TaxAuthorityName"),
        "effective_year": pick("EffectiveYear", "effective_year"),
        "status": pick("Status", "status"),
    }
    return {k: v for k, v in m.items() if v}


def sort_line_code(code: str) -> tuple[int, str]:
    """Sort key so numeric codes order first, ASSRFEES last, other codes between."""
    c = strip_field(code).upper()
    if c == "ASSRFEES":
        return (2, c)
    if c.isdigit():
        return (0, c.zfill(4))
    return (1, c)


def _mart_status_rank(status: str) -> int:
    """Rank mart row status by trust: active first, inactive last."""
    st = strip_field(status).upper()
    if st == "A":
        return 0
    if st == "":
        return 1
    if st == "I":
        return 2
    return 3


def _effective_year_number(raw: str | None) -> int:
    """Effective year as int for tie-breaks; unknown values sort last."""
    t = strip_field(raw)
    if t.isdigit():
        return int(t)
    return -1


def collapse_mart_tag_lines(lines: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """
    Deduplicate repeated mart levy rows within one TAG.

    County exports can include both active and inactive versions of the same
    authority/code pair. Keep one canonical row per pair:
      1) prefer status A over blank over I/other
      2) if tied, keep the newest effectiveYear
      3) if still tied, keep the first seen row
    """
    best_by_key: dict[tuple[str, str], dict[str, Any]] = {}
    best_rank: dict[tuple[str, str], tuple[int, int, int]] = {}
    for idx, ln in enumerate(lines):
        key = (
            strip_field(str(ln.get("code") or "")).upper(),
            strip_field(str(ln.get("authorityNameUpper") or "")).upper(),
        )
        rank = (
            _mart_status_rank(str(ln.get("status") or "")),
            -_effective_year_number(str(ln.get("effectiveYear") or "")),
            idx,
        )
        prev = best_rank.get(key)
        if prev is None or rank < prev:
            best_rank[key] = rank
            best_by_key[key] = ln
    return list(best_by_key.values())


def normalize_for_match(name: str) -> str:
    """
    Shared normalization for mart authority labels and DOLA legal names before
    fuzz.token_sort_ratio. Keeps both sides comparable so common abbreviations
    do not tank scores (e.g. METRO vs METROPOLITAN, DIST vs DISTRICT).

    Word-boundary rules use \\b so we do not alter METROPOLITAN (METRO is not a
    standalone token inside that word).
    """
    s = strip_field(name).upper()
    s = s.replace("&", " AND ")
    s = s.replace("#", " ")
    s = re.sub(r"[^A-Z0-9\s]", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    s = re.sub(r"\bDIST\b", "DISTRICT", s)
    s = re.sub(r"\bSCH\b", "SCHOOL", s)
    s = re.sub(r"\bVLG\b", "VILLAGE", s)
    s = re.sub(r"\bMD\b", "METROPOLITAN DISTRICT", s)
    s = re.sub(r"\bS SUBURBAN\b", "SOUTH SUBURBAN", s)
    # Typo occasionally seen in exports
    s = s.replace("DISTRRICT", "DISTRICT")
    # Mart lines often say METRO; DOLA legal names say METROPOLITAN
    s = re.sub(r"\bMETRO\b", "METROPOLITAN", s)
    return s


def load_overrides(path: Path) -> dict[str, dict[str, Any]]:
    """Load DOLA authority overrides keyed by uppercase mart authority label."""
    if not path.is_file():
        return {}
    data = json.loads(path.read_text(encoding="utf-8"))
    raw = data.get("byAuthorityUpper") or {}
    return {strip_field(k).upper(): v for k, v in raw.items()}


def build_entities_by_te_id(entities: list[dict[str, Any]]) -> dict[str, dict[str, Any]]:
    """One row per Tax Entity ID (Arapahoe-only list from load_dola_entities)."""
    out: dict[str, dict[str, Any]] = {}
    for e in entities:
        tid = _te_id_str(e.get("taxEntityId"))
        if not tid:
            continue
        if tid in out:
            print(
                f"Duplicate Tax Entity ID in DOLA Arapahoe rows: {tid} (keeping first).",
                file=sys.stderr,
            )
            continue
        out[tid] = e
    return out


def attach_levy_mills(
    dola: dict[str, Any],
    authority_upper: str,
    entities_by_te_id: dict[str, dict[str, Any]],
) -> dict[str, Any]:
    """
    Attach mills from the matched DOLA row when bond/debt purpose aligns with the
    mart label and matched legal name. Otherwise set mills to null (wrong is worse than missing).
    """
    method = dola.get("method")
    if method in ("none", "skipped"):
        return dola
    te_str = _te_id_str(dola.get("taxEntityId"))
    if not te_str:
        out = dict(dola)
        out["mills"] = None
        return out
    ent = entities_by_te_id.get(te_str)
    matched_name = strip_field(str(dola.get("matchedLegalName") or ""))
    if not ent:
        out = dict(dola)
        out["mills"] = None
        return out
    legal = strip_field(str(ent.get("legalName") or ""))

    mart_bond = mart_line_looks_bond_purpose(authority_upper)
    dola_bond = dola_name_looks_bond_purpose(matched_name or legal)

    if mart_bond != dola_bond:
        out = dict(dola)
        out["mills"] = None
        out["millsReason"] = "bond_purpose_mismatch"
        return out

    levy = ent.get("levyMills")
    if levy is None:
        out = dict(dola)
        out["mills"] = None
        return out

    out = dict(dola)
    out["mills"] = round(float(levy), 6)
    return out


def enrich_overrides_from_entities(
    overrides: dict[str, dict[str, Any]],
    entities: list[dict[str, Any]],
    min_score: float = 0.88,
) -> dict[str, dict[str, Any]]:
    """
    Fill taxEntityId / lgId on override rows by matching legalName to DOLA export
    rows (token_sort_ratio). Does not overwrite existing taxEntityId.
    """
    out: dict[str, dict[str, Any]] = {}
    resolved = 0
    for key, v in overrides.items():
        ovr = dict(v)
        if ovr.get("taxEntityId"):
            out[key] = ovr
            continue
        ln = ovr.get("legalName")
        if not ln or not entities:
            out[key] = ovr
            continue
        q = normalize_for_match(str(ln))
        best: dict[str, Any] | None = None
        best_score = -1.0
        for e in entities:
            s = fuzz.token_sort_ratio(q, e["norm"]) / 100.0
            if s > best_score:
                best_score = s
                best = e
        if best and best_score >= min_score:
            ovr["taxEntityId"] = best.get("taxEntityId")
            ovr["lgId"] = best.get("lgId")
            ovr["resolvedFromXlsx"] = True
            ovr["resolvedScore"] = round(best_score, 4)
            resolved += 1
        out[key] = ovr
    if resolved:
        print(f"Enriched {resolved} override rows from DOLA export (>= {min_score:.0%} name match).", file=sys.stderr)
    return out


def _dola_column_indices(headers: list[str]) -> tuple[int | None, int | None, int | None, int | None, int | None] | None:
    """Heuristic column detection for DOLA LGIS exports (xlsx or CSV). Returns None if no name column."""
    idx_name = None
    idx_entity = None
    idx_lgid = None
    idx_county = None
    idx_levy = None
    for i, h in enumerate(headers):
        hl = h.lower()
        if idx_name is None and "name" in hl and "tax" in hl:
            idx_name = i
        if idx_name is None and hl in ("tax entity name", "entity name", "legal name"):
            idx_name = i
        if idx_entity is None and "tax entity" in hl and "id" in hl:
            idx_entity = i
        if idx_lgid is None and ("lgid" in hl.replace(" ", "") or hl == "lg id"):
            idx_lgid = i
        if idx_county is None and "certifying" in hl and "county" in hl:
            idx_county = i
        if idx_levy is None and "levy" in hl and ("total" in hl or "budget" in hl):
            idx_levy = i

    if idx_name is None:
        for i, h in enumerate(headers):
            if "name" in h.lower() and "county" not in h.lower():
                idx_name = i
                break

    if idx_name is None:
        return None
    return idx_name, idx_entity, idx_lgid, idx_county, idx_levy


def _entities_from_dola_table_rows(
    rows: Any,
    idx_name: int,
    idx_entity: int | None,
    idx_lgid: int | None,
    idx_county: int | None,
    idx_levy: int | None,
    certifying_county_upper: str,
) -> list[dict[str, Any]]:
    """Build normalized entity dicts from DOLA table rows filtered to one certifying county."""
    entities: list[dict[str, Any]] = []
    for row in rows:
        if row is None:
            continue
        cells = list(row)
        if not cells or idx_name >= len(cells):
            continue
        if idx_county is not None and idx_county < len(cells):
            cty = strip_field(str(cells[idx_county] if cells[idx_county] is not None else ""))
            if cty.upper() != certifying_county_upper:
                continue
        legal = cells[idx_name]
        if legal is None or strip_field(str(legal)) == "":
            continue
        legal_s = strip_field(str(legal))
        te_id = ""
        if idx_entity is not None and idx_entity < len(cells) and cells[idx_entity] is not None:
            te_id = strip_field(str(cells[idx_entity]))
        lg = ""
        if idx_lgid is not None and idx_lgid < len(cells) and cells[idx_lgid] is not None:
            lg = strip_field(str(cells[idx_lgid]))
        levy_mills: float | None = None
        if idx_levy is not None and idx_levy < len(cells):
            levy_mills = parse_levy_mills_cell(cells[idx_levy])
        norm = normalize_for_match(legal_s)
        if not norm:
            continue
        entities.append(
            {
                "legalName": legal_s,
                "norm": norm,
                "taxEntityId": te_id or None,
                "lgId": lg or None,
                "levyMills": levy_mills,
            }
        )
    return entities


def load_dola_entities_csv(
    csv_path: Path, certifying_county: str = "Arapahoe"
) -> tuple[list[dict[str, Any]], str | None, bool]:
    """
    Parse DOLA Property Tax Entities export from CSV.

    Returns (entities, levy_header_or_none, county_filter_applied). The third value is
    True only when a Certifying County column exists and rows are restricted to ccu.
    """
    ccu = strip_field(certifying_county).upper()
    if not ccu:
        print("DOLA CSV: empty certifying county; skipping DOLA join.", file=sys.stderr)
        return [], None, False
    with csv_path.open(newline="", encoding="utf-8-sig", errors="replace") as f:
        reader = csv.reader(f)
        try:
            header_row = next(reader)
        except StopIteration:
            return [], None, False
        headers = [strip_field(str(h) if h is not None else "") for h in header_row]
        parsed = _dola_column_indices(headers)
        if parsed is None:
            print("Could not detect name column in DOLA CSV; skipping DOLA join.", file=sys.stderr)
            return [], None, False
        idx_name, idx_entity, idx_lgid, idx_county, idx_levy = parsed
        county_filter_applied = idx_county is not None
        if idx_county is None:
            print(
                "DOLA CSV: no Certifying County column; using all rows (may duplicate Tax Entity IDs across counties).",
                file=sys.stderr,
            )
        levy_header = headers[idx_levy] if idx_levy is not None else None
        entities = _entities_from_dola_table_rows(
            reader, idx_name, idx_entity, idx_lgid, idx_county, idx_levy, ccu
        )
        return entities, levy_header, county_filter_applied


def load_dola_entities_xlsx(
    xlsx_path: Path, certifying_county: str = "Arapahoe"
) -> tuple[list[dict[str, Any]], str | None, bool]:
    """Parse DOLA export from xlsx via openpyxl; returns (entities, levy header name, county_filter_applied)."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        print("openpyxl required for DOLA xlsx; skipping DOLA join.", file=sys.stderr)
        return [], None, False

    ccu = strip_field(certifying_county).upper()
    if not ccu:
        print("DOLA xlsx: empty certifying county; skipping DOLA join.", file=sys.stderr)
        return [], None, False
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    try:
        ws = wb.active
        rows = ws.iter_rows(values_only=True)
        try:
            header_row = next(rows)
        except StopIteration:
            return [], None, False
        headers = [strip_field(str(h) if h is not None else "") for h in header_row]
        parsed = _dola_column_indices(headers)
        if parsed is None:
            print("Could not detect name column in DOLA xlsx; skipping DOLA join.", file=sys.stderr)
            return [], None, False
        idx_name, idx_entity, idx_lgid, idx_county, idx_levy = parsed
        county_filter_applied = idx_county is not None
        if idx_county is None:
            print(
                "DOLA xlsx: no Certifying County column; using all rows (may duplicate Tax Entity IDs across counties).",
                file=sys.stderr,
            )
        levy_header = headers[idx_levy] if idx_levy is not None else None
        entities = _entities_from_dola_table_rows(
            rows, idx_name, idx_entity, idx_lgid, idx_county, idx_levy, ccu
        )
        return entities, levy_header, county_filter_applied
    finally:
        wb.close()


def load_dola_entities(
    path: Path, certifying_county: str = "Arapahoe"
) -> tuple[list[dict[str, Any]], str | None, bool]:
    """
    Load DOLA Tax Entity rows for one certifying county (avoids duplicate TE IDs across
    certifying counties when the export includes a county column). Attaches levyMills from
    the export's total levy column when present.
    Accepts .csv (UTF-8, optional BOM) or .xlsx. Returns (entities, levy_column_header_or_none, county_filter_applied).
    """
    if not path.is_file():
        return [], None, False
    suf = path.suffix.lower()
    if suf == ".csv":
        return load_dola_entities_csv(path, certifying_county)
    if suf in (".xlsx", ".xlsm"):
        return load_dola_entities_xlsx(path, certifying_county)
    print(f"Unsupported DOLA export format (expected .csv or .xlsx): {path}", file=sys.stderr)
    return [], None, False


def dola_match_for_mart_line(
    line_code: str,
    authority_upper: str,
    entities: list[dict[str, Any]],
    overrides: dict[str, dict[str, Any]],
    entities_by_te_id: dict[str, dict[str, Any]],
) -> dict[str, Any]:
    """ASSRFEES is an assessor fee row, not a taxing district — skip DOLA fuzzy matching."""
    code = strip_field(line_code).upper()
    if code == "ASSRFEES":
        return {
            "method": "skipped",
            "confidence": "high",
            "skipReason": "assessor_fee",
            "taxEntityId": None,
            "lgId": None,
            "matchedLegalName": None,
            "score": None,
        }
    au = strip_field(authority_upper).upper()
    ovr = overrides.get(au)
    dola = match_dola_line(authority_upper, entities, overrides)
    result = attach_levy_mills(dola, authority_upper, entities_by_te_id)
    if ovr is not None and ovr.get("millsOverride") is not None:
        try:
            mv = float(ovr["millsOverride"])
        except (TypeError, ValueError):
            return result
        out = dict(result)
        prev = out.get("mills")
        if isinstance(prev, (int, float)):
            out["dolaMills"] = round(float(prev), 6)
        out["mills"] = round(mv, 6)
        out["millsReason"] = "county_levy_table_override"
        return out
    return result


def _te_id_str(x: Any) -> str:
    """String form of a tax entity id, stripped, or empty when missing."""
    if x is None:
        return ""
    return strip_field(str(x))


def _coalesce_lg_id_from_entity(lg_id: Any, tax_entity_id: Any) -> Any:
    """Property tax export sometimes omits lgId; tax entity id is often '{lgId}/1'."""
    lid = strip_field(str(lg_id)) if lg_id is not None else ""
    if lid:
        return lid
    te = _te_id_str(tax_entity_id)
    if "/" in te:
        left = te.split("/", 1)[0].strip()
        if left.isdigit():
            return left
    return None


def match_dola_line(
    authority_upper: str,
    entities: list[dict[str, Any]],
    overrides: dict[str, dict[str, Any]],
) -> dict[str, Any]:
    """Resolve one mart authority line to a DOLA tax entity (override, fuzzy, or none)."""
    au = strip_field(authority_upper).upper()
    ovr = overrides.get(au)
    query_base = normalize_for_match(authority_upper)
    if ovr and ovr.get("legalName"):
        query_base = normalize_for_match(str(ovr["legalName"]))

    def ura_extra() -> dict[str, Any]:
        """Optional flag when an override marks an urban renewal context."""
        if ovr and ovr.get("ura"):
            return {"uraHint": True}
        return {}

    # Direct Tax Entity ID (after enrich from DOLA export) — exact row match
    if ovr and ovr.get("taxEntityId") and entities:
        want = _te_id_str(ovr.get("taxEntityId"))
        if want:
            for e in entities:
                if _te_id_str(e.get("taxEntityId")) == want:
                    return {
                        "method": "override",
                        "confidence": "high",
                        "taxEntityId": e.get("taxEntityId"),
                        "lgId": _coalesce_lg_id_from_entity(
                            e.get("lgId"), e.get("taxEntityId")
                        ),
                        "matchedLegalName": e.get("legalName"),
                        "score": 1.0,
                        **ura_extra(),
                    }

    if not entities:
        if ovr:
            return {
                "method": "override",
                "confidence": "low",
                "taxEntityId": ovr.get("taxEntityId"),
                "lgId": _coalesce_lg_id_from_entity(
                    ovr.get("lgId"), ovr.get("taxEntityId")
                ),
                "matchedLegalName": ovr.get("legalName"),
                "score": None,
                **ura_extra(),
            }
        return {"method": "none", "confidence": "low", "matchedLegalName": None, "score": None}

    best: dict[str, Any] | None = None
    best_score = -1.0
    for e in entities:
        score = fuzz.token_sort_ratio(query_base, e["norm"]) / 100.0
        if score > best_score:
            best_score = score
            best = e

    assert best is not None
    if best_score >= 0.92:
        conf = "high"
    elif best_score >= 0.78:
        conf = "medium"
    else:
        conf = "low"

    if ovr and ovr.get("legalName") and best_score < 0.78:
        query2 = normalize_for_match(str(ovr["legalName"]))
        for e in entities:
            score = fuzz.token_sort_ratio(query2, e["norm"]) / 100.0
            if score > best_score:
                best_score = score
                best = e
        if best_score >= 0.92:
            conf = "high"
        elif best_score >= 0.78:
            conf = "medium"
        else:
            conf = "low"

    if best_score < 0.70:
        return {
            "method": "none",
            "confidence": "low",
            "taxEntityId": None,
            "lgId": None,
            "matchedLegalName": None,
            "score": round(best_score, 4),
            **ura_extra(),
        }

    method = "fuzzy"
    if ovr:
        method = "override"

    return {
        "method": method,
        "confidence": conf,
        "taxEntityId": best.get("taxEntityId"),
        "lgId": _coalesce_lg_id_from_entity(
            best.get("lgId"), best.get("taxEntityId")
        ),
        "matchedLegalName": best.get("legalName"),
        "score": round(best_score, 4),
        **ura_extra(),
    }


def read_mart_groups(path: Path) -> tuple[dict[str, list[dict[str, Any]]], str]:
    """Read Mart_TA_TAG CSV into TAGId -> levy line rows plus tax year string."""
    with path.open(newline="", encoding="utf-8", errors="replace") as f:
        r = csv.DictReader(f)
        maps = mart_row_maps(r.fieldnames)
        if not maps.get("tag_id") or not maps.get("line_code") or not maps.get("authority_name"):
            raise SystemExit(f"Unexpected mart CSV headers: {r.fieldnames}")

        by_tag: dict[str, list[dict[str, Any]]] = defaultdict(list)
        tax_year = ""
        for row in r:
            tag = strip_field(row.get(maps["tag_id"], ""))
            if not tag:
                continue
            if not tax_year and maps.get("tax_year"):
                tax_year = strip_field(row.get(maps["tax_year"], ""))
            code = strip_field(row.get(maps["line_code"], ""))
            name = strip_field(row.get(maps["authority_name"], ""))
            eff = (
                strip_field(row.get(maps["effective_year"], ""))
                if maps.get("effective_year")
                else ""
            )
            st = strip_field(row.get(maps["status"], "")) if maps.get("status") else ""
            by_tag[tag].append(
                {
                    "code": code,
                    "authorityName": name,
                    "authorityNameUpper": name.upper(),
                    "effectiveYear": eff,
                    "status": st,
                }
            )
    return by_tag, tax_year or ""


# Mart LegalDescr often prefixes human text with subdivision/block/lot keys (~83% of rows).
_LEGAL_DESCR_PREFIX_RE = re.compile(
    r"^SubdivisionCd\s+\S+\s+SubdivisionName\s+.+?\s+Block\s+\S+\s+Lot\s+\S+\s+",
    re.IGNORECASE,
)


def legal_descr_display_tail(full: str) -> str:
    """Strip mart subdivision/block/lot prefix when present; else return trimmed full string."""
    s = strip_field(full)
    if not s:
        return ""
    m = _LEGAL_DESCR_PREFIX_RE.match(s)
    if m:
        tail = s[m.end() :].strip()
        return tail if tail else s
    return s


_LEGAL_DESCR_TYPE_RANK = (
    "MetesBounds",
    "UnPlatted",
    "PersonalDescr",
    "Platted",
)


def _legal_descr_type_rank(descr_type: str) -> int:
    t = strip_field(descr_type)
    try:
        return _LEGAL_DESCR_TYPE_RANK.index(t)
    except ValueError:
        return len(_LEGAL_DESCR_TYPE_RANK)


def pick_legal_description_display(rows: list[tuple[str, str]]) -> str | None:
    """
    Prefer MetesBounds / UnPlatted / PersonalDescr DisplayDescr from Mart_DescrHeader.
    Platted-only rows often repeat mart keys; keep Main Parcel display for those.
    """
    preferred = frozenset({"MetesBounds", "UnPlatted", "PersonalDescr"})
    sorted_rows = sorted(rows, key=lambda item: _legal_descr_type_rank(item[0]))
    for dtype, display in sorted_rows:
        if dtype not in preferred:
            continue
        tail = legal_descr_display_tail(display)
        if tail:
            return tail
    return None


def read_legal_description_display_by_pin(path: Path) -> dict[str, str]:
    """PIN -> display-friendly legal text from Mart_DescrHeader when a preferred row exists."""
    if not path.is_file():
        return {}
    by_pin: dict[str, list[tuple[str, str]]] = defaultdict(list)
    with path.open(newline="", encoding="utf-8", errors="replace") as f:
        for row in csv.DictReader(f):
            pin = normalize_pin(strip_field(row.get("PIN") or row.get("Pin") or ""))
            display = strip_field(row.get("DisplayDescr", ""))
            dtype = strip_field(row.get("DescrType", ""))
            if not pin or not display:
                continue
            by_pin[pin].append((dtype, display))
    out: dict[str, str] = {}
    for pin, items in by_pin.items():
        picked = pick_legal_description_display(items)
        if picked:
            out[pin] = picked
    return out


def ownership_type_label_from_owner_lp_types(lp_types: list[str]) -> str | None:
    """County-style Ownership Type from Mart_LegalParty owner rows (LPRType=Owner).

    Vesting labels such as Joint Tenancy are not exported in the mart; when every owner
    row is Individual we match the common county parcel-page label for co-owners.
    """
    types = [strip_field(t) for t in lp_types if strip_field(t)]
    if not types:
        return None
    if len(types) == 1:
        return types[0]
    unique = list(dict.fromkeys(types))
    if len(unique) == 1 and unique[0] == "Individual":
        return "Joint Tenancy"
    return ", ".join(unique)


def read_ownership_type_by_pin(path: Path) -> dict[str, str]:
    """PIN -> Ownership Type label from Mart_LegalParty owner rows."""
    if not path.is_file():
        return {}
    by_pin: dict[str, list[str]] = defaultdict(list)
    with path.open(newline="", encoding="utf-8", errors="replace") as f:
        for row in csv.DictReader(f):
            pin = normalize_pin(strip_field(row.get("PIN") or row.get("Pin") or ""))
            lpr_type = strip_field(row.get("LPRType", ""))
            if lpr_type.upper() != "OWNER":
                continue
            lp_type = strip_field(row.get("LPType", ""))
            if not pin or not lp_type:
                continue
            by_pin[pin].append(lp_type)
    out: dict[str, str] = {}
    for pin, lp_types in by_pin.items():
        label = ownership_type_label_from_owner_lp_types(lp_types)
        if label:
            out[pin] = label
    return out


def parse_acreage_cell(val: Any) -> float | None:
    """Parse Mart_RDE_LndAll Acreage; returns None if missing or invalid."""
    return parse_parcel_value_cell(val)


def format_acreage_display(total: float) -> str:
    """County parcel record shows acreage to four decimal places (e.g. 0.0540)."""
    if not math.isfinite(total):
        return ""
    return f"{total:.4f}"


def format_land_units_display(row: dict[str, str]) -> str:
    """County Land Line Units column (e.g. 1.0000 LT from Uts + UnitTp)."""
    uts = strip_field(row.get("Uts", ""))
    unit_tp = strip_field(row.get("UnitTp", ""))
    if not uts:
        return ""
    try:
        units = f"{float(uts):.4f}"
    except ValueError:
        units = uts
    return f"{units} {unit_tp}".strip() if unit_tp else units


def land_table_row_from_csv(row: dict[str, str]) -> dict[str, str] | None:
    """One Land Line table row: Units + land-use description (not top-level Land Use)."""
    units = format_land_units_display(row)
    land_use = strip_field(row.get("UseCdDscr", ""))
    if not units and not land_use:
        return None
    out: dict[str, str] = {}
    if units:
        out["units"] = units
    if land_use:
        out["landUse"] = land_use
    return out


def format_county_count(val: Any) -> str:
    """Bed/bath-style counts as on county parcel record (e.g. 3.00)."""
    s = strip_field(str(val)) if val is not None else ""
    if not s:
        return ""
    try:
        return f"{float(s):.2f}"
    except ValueError:
        return s


def format_county_sqft(val: Any) -> str:
    s = strip_field(str(val)) if val is not None else ""
    if not s:
        return ""
    try:
        n = float(s)
        if n == int(n):
            return str(int(n))
        return s
    except ValueError:
        return s


# County PPINum.aspx attribute labels and Mart_RDE_BLD columns (stable order).
# Fireplaces appears on the county page between Roof and Exterior Wall but is not
# exported in Mart_RDE_BLD / Mart_RDE_Xfob — keep the slot for county order (empty).
BUILDING_ATTRIBUTE_FIELDS: tuple[tuple[str, str | None], ...] = (
    ("Quality Grade", "QualityCd"),
    ("Improvement Type", "ImprTpDscr"),
    ("Bedrooms", "BedCount"),
    ("Bathrooms", "BathCount"),
    ("Architectural", "ImprMdlCdDscr"),
    ("Heat Method", "HeatCd1"),
    ("Cool Method", "CoolCd1"),
    ("Year Built", "ActYear"),
    ("Roof", "RoofCd1"),
    ("Fireplaces", None),
    ("Exterior Wall", "ExtwallCd1"),
    ("Construction Type", "Class"),
)

COUNTY_DECIMAL_ATTRIBUTE_KEYS = frozenset({"BedCount", "BathCount"})

# Always emit these labels (even empty) so county attribute order stays stable.
ALWAYS_EMIT_BUILDING_ATTRIBUTE_LABELS = frozenset({"Fireplaces"})

PERMIT_STATUS_LABELS: dict[str, str] = {
    "P": "Pending",
    "C": "Complete",
    "A": "Assigned",
    "S": "Submitted",
    "V": "Void",
    "M": "Incomplete",
}


def building_record_from_csv(row: dict[str, str]) -> dict[str, Any] | None:
    """Structured building block matching county attributes + area tables."""
    building_num = strip_field(row.get("num", ""))
    attributes: list[dict[str, str]] = []
    for label, key in BUILDING_ATTRIBUTE_FIELDS:
        if key is None:
            if label in ALWAYS_EMIT_BUILDING_ATTRIBUTE_LABELS:
                attributes.append({"label": label, "value": ""})
            continue
        raw = strip_field(row.get(key, ""))
        if not raw:
            continue
        if key in COUNTY_DECIMAL_ATTRIBUTE_KEYS:
            value = format_county_count(raw)
        else:
            value = raw
        attributes.append({"label": label, "value": value})
    areas: list[dict[str, str]] = []
    for i in range(1, 20):
        descr = strip_field(row.get(f"SarCatDscr{i}", ""))
        if not descr:
            continue
        area_raw = strip_field(row.get(f"SarCatArea{i}", ""))
        areas.append(
            {
                "description": descr,
                "sqFt": format_county_sqft(area_raw) if area_raw else "",
            }
        )
    total_area = format_county_sqft(row.get("BaseArea"))
    if not building_num and not attributes and not areas:
        return None
    rec: dict[str, Any] = {"buildingNum": building_num or "1"}
    if attributes:
        rec["attributes"] = attributes
    if areas:
        rec["areas"] = areas
    if total_area:
        rec["totalArea"] = total_area
    return rec


def read_land_fields_by_pin(path: Path) -> dict[str, dict[str, Any]]:
    """PIN -> acreage and landLines table rows from Mart_RDE_LndAll."""
    if not path.is_file():
        return {}
    by_pin: dict[str, list[dict[str, str]]] = defaultdict(list)
    with path.open(newline="", encoding="utf-8", errors="replace") as f:
        for row in csv.DictReader(f):
            pin = normalize_pin(strip_field(row.get("Pin") or row.get("PIN") or ""))
            if not pin:
                continue
            by_pin[pin].append(row)
    out: dict[str, dict[str, Any]] = {}
    for pin, rows in by_pin.items():
        rows_sorted = sorted(
            rows,
            key=lambda r: int(strip_field(r.get("Num", "")) or "0"),
        )
        land_lines = [
            line
            for line in (land_table_row_from_csv(r) for r in rows_sorted)
            if line
        ]
        total_acre = 0.0
        has_acre = False
        for row in rows_sorted:
            acre = parse_acreage_cell(row.get("Acreage"))
            if acre is not None:
                total_acre += acre
                has_acre = True
        entry: dict[str, Any] = {}
        if has_acre:
            entry["acreage"] = format_acreage_display(total_acre)
        if land_lines:
            entry["landLines"] = land_lines
        if entry:
            out[pin] = entry
    return out


def read_building_fields_by_pin(
    bld_path: Path,
    xfob_path: Path,
) -> dict[str, dict[str, Any]]:
    """PIN -> landUse (ImprTpDscr) and buildings[] from Mart_RDE_BLD."""
    del xfob_path  # reserved for a future county-parity pass; not on PPINum layout today
    bld_by_pin: dict[str, list[dict[str, str]]] = defaultdict(list)
    if bld_path.is_file():
        with bld_path.open(newline="", encoding="utf-8", errors="replace") as f:
            for row in csv.DictReader(f):
                pin = normalize_pin(strip_field(row.get("Pin") or row.get("PIN") or ""))
                if pin:
                    bld_by_pin[pin].append(row)
    out: dict[str, dict[str, Any]] = {}
    for pin, rows in bld_by_pin.items():
        bld_rows = sorted(
            rows,
            key=lambda r: int(strip_field(r.get("num", "")) or "0"),
        )
        buildings = [
            bld
            for bld in (building_record_from_csv(r) for r in bld_rows)
            if bld
        ]
        entry: dict[str, Any] = {}
        for row in bld_rows:
            typ = strip_field(row.get("ImprTpDscr", ""))
            if typ:
                entry["landUse"] = typ
                break
        if buildings:
            entry["buildings"] = buildings
        if entry:
            out[pin] = entry
    return out


def format_county_mm_dd_yyyy(raw: str) -> str:
    """Mart YYYYMMDD → county parcel-page date (MM-DD-YYYY)."""
    s = strip_field(raw)
    if len(s) == 8 and s.isdigit():
        return f"{s[4:6]}-{s[6:8]}-{s[0:4]}"
    return s


def normalize_state_use_cd(raw: str) -> str:
    """Normalize StateUseCd for xlsx lookup (strip float suffix; pad 3-digit codes)."""
    s = strip_field(raw)
    if not s:
        return ""
    try:
        n = float(s)
        if n == int(n):
            s = str(int(n))
    except ValueError:
        pass
    if s.isdigit() and len(s) == 3:
        return s.zfill(4)
    return s


def format_book_page_display(book: str, page: str) -> str:
    """County Sale 'Book Page' cell (book + page with a space)."""
    b = strip_field(book)
    p = strip_field(page)
    if not b and not p:
        return ""
    return f"{b} {p}".strip()


def transfer_sale_row_from_csv(row: dict[str, str]) -> dict[str, Any] | None:
    """One Sale history row when Book+Page are present (matches PPINum sale table)."""
    book = strip_field(row.get("Book", ""))
    page = strip_field(row.get("Page", ""))
    if not book or not page:
        return None
    date = format_county_mm_dd_yyyy(row.get("DocDate", ""))
    price = parse_parcel_value_cell(row.get("Consid"))
    out: dict[str, Any] = {
        "bookPage": format_book_page_display(book, page),
        "date": date or "",
        "sortDate": strip_field(row.get("DocDate", "")),
    }
    if price is not None:
        out["price"] = price
    # County PPINum "Type" column is typically blank; omit rather than invent labels.
    return out


def read_transfers_by_pin(path: Path) -> dict[str, list[dict[str, Any]]]:
    """PIN -> Sale history rows from Mart_Transfers (Book+Page only; newest first)."""
    if not path.is_file():
        return {}
    by_pin: dict[str, list[dict[str, Any]]] = defaultdict(list)
    with path.open(newline="", encoding="utf-8", errors="replace") as f:
        for row in csv.DictReader(f):
            pin = normalize_pin(strip_field(row.get("PIN") or row.get("Pin") or ""))
            if not pin:
                continue
            sale = transfer_sale_row_from_csv(row)
            if sale:
                by_pin[pin].append(sale)
    out: dict[str, list[dict[str, Any]]] = {}
    for pin, rows in by_pin.items():
        rows_sorted = sorted(
            rows,
            key=lambda r: strip_field(str(r.get("sortDate", ""))),
            reverse=True,
        )
        cleaned: list[dict[str, Any]] = []
        for r in rows_sorted:
            item = {k: v for k, v in r.items() if k != "sortDate"}
            cleaned.append(item)
        out[pin] = cleaned
    return out


def permit_row_from_csv(row: dict[str, str]) -> dict[str, Any] | None:
    """One permit row for the extended parcel-record table."""
    permit_num = strip_field(row.get("Permit_Num", ""))
    dscr = strip_field(row.get("Dscr", ""))
    status_raw = strip_field(row.get("Status", ""))
    issue_raw = strip_field(row.get("Issue_Dt", ""))
    final_raw = strip_field(row.get("Final_Dt", ""))
    est_val = parse_parcel_value_cell(row.get("Est_Val"))
    if not permit_num and not dscr and est_val is None:
        return None
    status = PERMIT_STATUS_LABELS.get(status_raw.upper(), status_raw)
    issue = format_county_mm_dd_yyyy(issue_raw) if issue_raw and issue_raw != "18991230" else ""
    final = format_county_mm_dd_yyyy(final_raw) if final_raw and final_raw != "18991230" else ""
    out: dict[str, Any] = {
        "sortDate": issue_raw if issue_raw and issue_raw != "18991230" else "",
    }
    if permit_num:
        out["permitNum"] = permit_num
    if status:
        out["status"] = status
    if dscr:
        out["description"] = dscr
    if issue:
        out["issueDate"] = issue
    if final:
        out["finalDate"] = final
    if est_val is not None:
        out["estimatedValue"] = est_val
    return out


def read_permits_by_pin(path: Path) -> dict[str, list[dict[str, Any]]]:
    """PIN -> permit rows from Mart_RDE_Permit (newest issue date first)."""
    if not path.is_file():
        return {}
    by_pin: dict[str, list[dict[str, Any]]] = defaultdict(list)
    with path.open(newline="", encoding="utf-8", errors="replace") as f:
        for row in csv.DictReader(f):
            pin = normalize_pin(strip_field(row.get("PIN") or row.get("Pin") or ""))
            if not pin:
                continue
            permit = permit_row_from_csv(row)
            if permit:
                by_pin[pin].append(permit)
    out: dict[str, list[dict[str, Any]]] = {}
    for pin, rows in by_pin.items():
        rows_sorted = sorted(
            rows,
            key=lambda r: (
                strip_field(str(r.get("sortDate", ""))),
                strip_field(str(r.get("permitNum", ""))),
            ),
            reverse=True,
        )
        cleaned: list[dict[str, Any]] = []
        for r in rows_sorted:
            item = {k: v for k, v in r.items() if k != "sortDate"}
            cleaned.append(item)
        out[pin] = cleaned
    return out


def read_xlsx_code_description_map(
    path: Path,
    *,
    code_col: int = 0,
    desc_col: int = 1,
) -> dict[str, str]:
    """Read a two-column code→description sheet (NBHD / State Class Codes xlsx)."""
    if not path.is_file():
        return {}
    try:
        from openpyxl import load_workbook
    except ImportError:
        print(
            f"openpyxl not installed; skipping lookup workbook {path.name}",
            file=sys.stderr,
        )
        return {}
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        out: dict[str, str] = {}
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if not row or len(row) <= max(code_col, desc_col):
                continue
            if i == 0:
                # Skip header row (Neighborhood Code / State Use Code).
                continue
            code_raw = row[code_col]
            desc_raw = row[desc_col]
            if code_raw is None:
                continue
            if isinstance(code_raw, (int, float)):
                code = str(int(code_raw)) if float(code_raw) == int(code_raw) else str(code_raw)
            else:
                code = normalize_state_use_cd(str(code_raw))
            if not code:
                continue
            desc = strip_field(str(desc_raw) if desc_raw is not None else "")
            if desc:
                out[code] = desc
        return out
    finally:
        wb.close()


def read_nbhd_description_by_code(path: Path) -> dict[str, str]:
    """NBHD codes xlsx → code → neighborhood name."""
    return read_xlsx_code_description_map(path)


def read_state_class_description_by_code(path: Path) -> dict[str, str]:
    """State Class Codes xlsx → StateUseCd → description (col 0 code, col 2 description)."""
    return read_xlsx_code_description_map(path, code_col=0, desc_col=2)


def attach_state_use_label(
    rec: dict[str, Any],
    state_class_by_code: dict[str, str],
) -> bool:
    """Set stateUseLabel from Main Parcel StateUseCd + State Class Codes xlsx."""
    code = normalize_state_use_cd(str(rec.get("stateUseCd") or ""))
    if not code:
        return False
    # Prefer normalized code on the record for display/debug.
    if code != strip_field(str(rec.get("stateUseCd") or "")):
        rec["stateUseCd"] = code
    label = state_class_by_code.get(code)
    if not label:
        return False
    rec["stateUseLabel"] = label
    return True


def enrich_parcel_record_from_sibling_marts(
    parcel_record_map: dict[str, dict[str, Any]],
    *,
    legal_descriptions_path: Path | None,
    legal_parties_path: Path | None,
    land_path: Path | None = None,
    building_path: Path | None = None,
    building_xfob_path: Path | None = None,
    transfers_path: Path | None = None,
    permits_path: Path | None = None,
    state_class_xlsx_path: Path | None = None,
    nbhd_xlsx_path: Path | None = None,
) -> dict[str, int]:
    """Merge Phase 2–4 sibling mart tables / lookup workbooks into parcel-record rows."""
    legal_by_pin = (
        read_legal_description_display_by_pin(legal_descriptions_path)
        if legal_descriptions_path and legal_descriptions_path.is_file()
        else {}
    )
    ownership_by_pin = (
        read_ownership_type_by_pin(legal_parties_path)
        if legal_parties_path and legal_parties_path.is_file()
        else {}
    )
    land_by_pin = read_land_fields_by_pin(land_path) if land_path else {}
    building_by_pin = (
        read_building_fields_by_pin(building_path, building_xfob_path)
        if building_path and building_xfob_path
        else {}
    )
    transfers_by_pin = (
        read_transfers_by_pin(transfers_path) if transfers_path else {}
    )
    permits_by_pin = read_permits_by_pin(permits_path) if permits_path else {}
    state_class_by_code = (
        read_state_class_description_by_code(state_class_xlsx_path)
        if state_class_xlsx_path
        else {}
    )
    # NBHD xlsx is loaded for readiness / future join only: Main Parcel has no
    # neighborhood-code column, and SubdivisionName→NBHD guesses disagree with
    # live PPINum (do not infer neighborhood from subdivision).
    if nbhd_xlsx_path and nbhd_xlsx_path.is_file():
        nbhd_count = len(read_nbhd_description_by_code(nbhd_xlsx_path))
        if nbhd_count:
            print(
                f"NBHD lookup loaded ({nbhd_count} codes); not joined - "
                "no neighborhood code on Main Parcel CSV",
                file=sys.stderr,
            )

    counts = {
        "legalDescrDisplay": 0,
        "ownershipType": 0,
        "land": 0,
        "building": 0,
        "transfers": 0,
        "permits": 0,
        "stateUseLabel": 0,
    }
    for pin, rec in parcel_record_map.items():
        mart_legal = legal_by_pin.get(pin)
        if mart_legal:
            rec["legalDescrDisplay"] = mart_legal
            counts["legalDescrDisplay"] += 1
        ownership = ownership_by_pin.get(pin)
        if ownership:
            rec["ownershipType"] = ownership
            counts["ownershipType"] += 1
        land = land_by_pin.get(pin)
        if land:
            rec.update(land)
            counts["land"] += 1
        building = building_by_pin.get(pin)
        if building:
            if building.get("landUse"):
                rec["landUse"] = building["landUse"]
            if building.get("buildings"):
                rec["buildings"] = building["buildings"]
            counts["building"] += 1
        transfers = transfers_by_pin.get(pin)
        if transfers:
            rec["transfers"] = transfers
            counts["transfers"] += 1
        permits = permits_by_pin.get(pin)
        if permits:
            rec["permits"] = permits
            counts["permits"] += 1
        if state_class_by_code and attach_state_use_label(rec, state_class_by_code):
            counts["stateUseLabel"] += 1
    return counts


def print_parcel_record_shard_size_stats(shard_dir: Path) -> None:
    """Log shard size distribution so joins can be checked for bloat before shipping."""
    sizes = sorted(p.stat().st_size for p in shard_dir.glob("*.json"))
    if not sizes:
        return
    total_mb = sum(sizes) / (1024 * 1024)
    p90 = sizes[int(len(sizes) * 0.9)]
    p99 = sizes[int(len(sizes) * 0.99)]
    over_500 = sum(1 for s in sizes if s > 500 * 1024)
    over_1m = sum(1 for s in sizes if s > 1024 * 1024)
    print(
        f"Shard size stats: {len(sizes)} files, {total_mb:.2f} MiB total; "
        f"min {sizes[0] / 1024:.1f} KiB, median {statistics.median(sizes) / 1024:.1f} KiB, "
        f"mean {statistics.mean(sizes) / 1024:.1f} KiB, max {sizes[-1] / 1024:.1f} KiB; "
        f"p90 {p90 / 1024:.1f} KiB, p99 {p99 / 1024:.1f} KiB; "
        f">{500} KiB: {over_500}, >1 MiB: {over_1m}",
        file=sys.stderr,
    )


def _optional_str(row: dict[str, str], key: str) -> str | None:
    t = strip_field(row.get(key, ""))
    return t if t else None


def read_parcel_record_map(path: Path) -> dict[str, dict[str, Any]]:
    """First Main Parcel row per PIN — extended county parcel record fields for lazy UI load."""
    _, _, parcel_record_map = read_main_parcel_maps(path)
    return parcel_record_map


def read_pin_map(path: Path) -> dict[str, dict[str, Any]]:
    """First Main Parcel row per PIN for tag and values; AIN may be filled from a later row if missing."""
    pin_map, _, _ = read_main_parcel_maps(path)
    return pin_map


# Keep in sync with PARCEL_RECORD_SHARD_PREFIX_LENGTH in arapahoeParcelLevyData.ts
PARCEL_RECORD_SHARD_PREFIX_LEN = 6


def write_parcel_record_shards(
    out_dir: Path,
    parcel_record_map: dict[str, dict[str, Any]],
    parcel_snapshot: dict[str, Any],
    *,
    separators: tuple[str, str],
) -> None:
    """Write plain JSON shards by PIN prefix (one small fetch per lookup)."""
    shard_dir = out_dir / "arapahoe-parcel-record-by-pin"
    if shard_dir.exists():
        for old in shard_dir.glob("*.json.gz"):
            old.unlink()
        for old in shard_dir.glob("*.json"):
            old.unlink()
    shard_dir.mkdir(parents=True, exist_ok=True)

    legacy_mono = out_dir / "arapahoe-parcel-record-by-pin.json.gz"
    if legacy_mono.exists():
        legacy_mono.unlink()
    legacy_mono_json = out_dir / "arapahoe-parcel-record-by-pin.json"
    if legacy_mono_json.exists():
        legacy_mono_json.unlink()

    shards: dict[str, dict[str, dict[str, Any]]] = defaultdict(dict)
    for pin, row in parcel_record_map.items():
        prefix = pin[:PARCEL_RECORD_SHARD_PREFIX_LEN]
        shards[prefix][pin] = row

    for prefix in sorted(shards):
        by_pin = shards[prefix]
        path = shard_dir / f"{prefix}.json"
        path.write_text(
            json.dumps(
                {
                    "snapshot": parcel_snapshot,
                    "pinDigits": 9,
                    "shardPrefix": prefix,
                    "byPin": by_pin,
                },
                separators=separators,
            ),
            encoding="utf-8",
        )

    total_bytes = sum((shard_dir / f"{prefix}.json").stat().st_size for prefix in shards)
    total_mb = total_bytes / (1024 * 1024)
    print(
        f"Wrote {shard_dir}/ ({len(shards)} shards, {len(parcel_record_map)} pins, "
        f"{total_mb:.2f} MiB total)",
        file=sys.stderr,
    )
    print_parcel_record_shard_size_stats(shard_dir)


def normalize_bundled_as_of(raw: str) -> str:
    o = raw.strip()
    if re.fullmatch(r"\d{4}-\d{2}-\d{2}", o):
        return f"{o}T12:00:00Z"
    return o


def read_county_data_as_of_file(county_mart_dir: Path) -> str:
    """YYYY-MM-DD from county-mart/data-as-of.txt (maintainer sets on mart download)."""
    path = county_mart_dir / COUNTY_DATA_AS_OF_FILE
    if not path.is_file():
        raise SystemExit(
            f"Missing {path}. Add one line YYYY-MM-DD (date you downloaded this mart batch). "
            "Or pass --bundled-as-of YYYY-MM-DD."
        )
    line = path.read_text(encoding="utf-8").strip().splitlines()[0].strip()
    if not re.fullmatch(r"\d{4}-\d{2}-\d{2}", line):
        raise SystemExit(
            f"{path} must be a single YYYY-MM-DD line; got: {line!r}"
        )
    return normalize_bundled_as_of(line)


def resolve_bundled_as_of(override: str | None, county_mart_dir: Path) -> str:
    """Read maintainer date from data-as-of.txt, or --bundled-as-of override."""
    if override:
        return normalize_bundled_as_of(override)
    return read_county_data_as_of_file(county_mart_dir)


def main() -> None:
    """CLI entry: read county exports, join DOLA, write public/data JSON artifacts."""
    ap = argparse.ArgumentParser(description="Build Arapahoe parcel levy index JSON.")
    ap.add_argument("--main-parcel", type=Path, default=DEFAULT_MAIN)
    ap.add_argument("--mart-ta-tag", type=Path, default=DEFAULT_MART)
    ap.add_argument(
        "--legal-descriptions",
        type=Path,
        default=DEFAULT_LEGAL_DESCRIPTIONS,
        help="Mart_DescrHeader CSV (Parcel Legal Descriptions). Optional; skipped if missing.",
    )
    ap.add_argument(
        "--legal-parties",
        type=Path,
        default=DEFAULT_LEGAL_PARTIES,
        help="Mart_LegalParty CSV (Parcel Legal Parties). Optional; skipped if missing.",
    )
    ap.add_argument(
        "--land",
        type=Path,
        default=DEFAULT_LAND,
        help="Mart_RDE_LndAll CSV (Parcel Land Information). Optional; skipped if missing.",
    )
    ap.add_argument(
        "--building",
        type=Path,
        default=DEFAULT_BUILDING,
        help="Mart_RDE_BLD CSV (Parcel Building Information). Optional; skipped if missing.",
    )
    ap.add_argument(
        "--building-xfob",
        type=Path,
        default=DEFAULT_BUILDING_XFOB,
        help="Mart_RDE_Xfob CSV (Parcel Building Extra Features). Optional; skipped if missing.",
    )
    ap.add_argument(
        "--transfers",
        type=Path,
        default=DEFAULT_TRANSFERS,
        help="Mart_Transfers CSV (Parcel Transfer Information). Optional; skipped if missing.",
    )
    ap.add_argument(
        "--permits",
        type=Path,
        default=DEFAULT_PERMITS,
        help="Mart_RDE_Permit CSV (Parcel Permit Information). Optional; skipped if missing.",
    )
    ap.add_argument(
        "--state-class-xlsx",
        type=Path,
        default=DEFAULT_STATE_CLASS_XLSX,
        help="State Class Codes xlsx (labels for Main Parcel StateUseCd). Optional.",
    )
    ap.add_argument(
        "--nbhd-xlsx",
        type=Path,
        default=DEFAULT_NBHD_XLSX,
        help="NBHD codes xlsx (lookup only today — Main Parcel has no neighborhood code column).",
    )
    ap.add_argument(
        "--dola-export",
        "--dola-xlsx",
        type=Path,
        default=None,
        dest="dola_export",
        metavar="PATH",
        help="DOLA LGIS Property Tax Entities export (.csv or .xlsx). "
        "Default: supporting-data/dola/property-tax-entities-export.csv if present, else .xlsx.",
    )
    ap.add_argument(
        "--dola-certifying-county",
        default="Arapahoe",
        help=(
            "Case-insensitive certifying county label for the DOLA export. "
            "If the file has a Certifying County column, only matching rows are used and "
            "the snapshot records dolaCertifyingCounty. "
            "If that column is missing, all rows are used and dolaCertifyingCounty is omitted."
        ),
    )
    ap.add_argument("--overrides", type=Path, default=DEFAULT_OVERRIDES)
    ap.add_argument("--out-dir", type=Path, default=DEFAULT_OUT_DIR)
    ap.add_argument("--skip-pin-map", action="store_true", help="Only emit stacks-by-tag-id JSON.")
    ap.add_argument(
        "--bundled-as-of",
        default=None,
        metavar="DATE",
        help=(
            "Override snapshot bundledAsOf (YYYY-MM-DD or ISO). "
            f"Default: {COUNTY_DATA_AS_OF_FILE} under county-mart/."
        ),
    )
    args = ap.parse_args()

    if not args.main_parcel.is_file():
        raise SystemExit(f"Missing main parcel CSV: {args.main_parcel}")
    if not args.mart_ta_tag.is_file():
        raise SystemExit(f"Missing mart CSV: {args.mart_ta_tag}")

    overrides = load_overrides(args.overrides)
    dola_path = args.dola_export if args.dola_export is not None else default_dola_export_path()
    dola_cc = strip_field(args.dola_certifying_county) or "Arapahoe"
    entities, levy_col_header, dola_county_filter_applied = load_dola_entities(
        dola_path, dola_cc
    )
    entities_by_te_id = build_entities_by_te_id(entities)
    if entities:
        if dola_county_filter_applied:
            print(
                f"DOLA entities loaded: {len(entities)} (certifying county {dola_cc} only)",
                file=sys.stderr,
            )
        else:
            print(f"DOLA entities loaded: {len(entities)}", file=sys.stderr)
        if levy_col_header:
            print(f"DOLA levy column: {levy_col_header}", file=sys.stderr)
    else:
        print("No DOLA export or empty parse; emitting matches as method=none.", file=sys.stderr)

    overrides = enrich_overrides_from_entities(overrides, entities)

    by_tag_raw, tax_year = read_mart_groups(args.mart_ta_tag)
    bundled_as_of = resolve_bundled_as_of(args.bundled_as_of, COUNTY_MART)
    print(f"Snapshot bundledAsOf: {bundled_as_of}", file=sys.stderr)

    stacks: dict[str, Any] = {}
    for tag_id, lines in by_tag_raw.items():
        lines_collapsed = collapse_mart_tag_lines(lines)
        lines_sorted = sorted(lines_collapsed, key=lambda x: sort_line_code(x["code"]))
        built_lines = []
        for ln in lines_sorted:
            dola = dola_match_for_mart_line(
                ln["code"], ln["authorityNameUpper"], entities, overrides, entities_by_te_id
            )
            built_lines.append(
                {
                    "code": ln["code"],
                    "authorityName": ln["authorityName"],
                    "effectiveYear": ln["effectiveYear"] or None,
                    "status": ln["status"] or None,
                    "dolaMatch": dola,
                }
            )
        stacks[tag_id] = {
            "tagId": tag_id,
            "taxYear": tax_year or None,
            "levyAspxUrl": f"{LEVY_ASPX_BASE}{tag_id}",
            "lines": built_lines,
        }

    snapshot = {
        "bundledAsOf": bundled_as_of,
        "source": "Arapahoe County datamart: Mart_TA_TAG + Main Parcel (Pin → TAGId, AIN, OwnerList, values)",
        "taxYear": tax_year or None,
        "dolaSource": str(dola_path.name) if dola_path.is_file() else None,
        "dolaRowCount": len(entities),
        "dolaCertifyingCounty": dola_cc if dola_county_filter_applied else None,
        "dolaLevyColumn": levy_col_header,
    }

    args.out_dir.mkdir(parents=True, exist_ok=True)
    sep = (",", ":")
    stacks_path = args.out_dir / "arapahoe-levy-stacks-by-tag-id.json"

    if args.skip_pin_map:
        stacks_path.write_text(
            json.dumps({"snapshot": snapshot, "stacksByTagId": stacks}, separators=sep),
            encoding="utf-8",
        )
        print(f"Wrote {stacks_path} ({len(stacks)} TAG stacks)", file=sys.stderr)
        print(
            "Skipping arapahoe-pin-to-tag.json, arapahoe-situs-to-pins.json, "
            "and arapahoe-parcel-record-by-pin shards (--skip-pin-map).",
            file=sys.stderr,
        )
        return

    pin_map, situs_map, parcel_record_map = read_main_parcel_maps(args.main_parcel)
    join_counts = enrich_parcel_record_from_sibling_marts(
        parcel_record_map,
        legal_descriptions_path=args.legal_descriptions,
        legal_parties_path=args.legal_parties,
        land_path=args.land,
        building_path=args.building,
        building_xfob_path=args.building_xfob,
        transfers_path=args.transfers,
        permits_path=args.permits,
        state_class_xlsx_path=args.state_class_xlsx,
        nbhd_xlsx_path=args.nbhd_xlsx,
    )
    if any(join_counts.values()):
        print(
            "Parcel record sibling joins: "
            + ", ".join(f"{k}={v}" for k, v in join_counts.items() if v),
            file=sys.stderr,
        )
    used_tag_ids = {v["tagId"] for v in pin_map.values()}
    stacks_out = {k: v for k, v in stacks.items() if k in used_tag_ids}
    if len(stacks_out) < len(stacks):
        print(
            f"Filtered stacks: {len(stacks)} -> {len(stacks_out)} (TAGIds on parcels only)",
            file=sys.stderr,
        )
    stacks_path.write_text(
        json.dumps({"snapshot": snapshot, "stacksByTagId": stacks_out}, separators=sep),
        encoding="utf-8",
    )
    print(f"Wrote {stacks_path} ({len(stacks_out)} TAG stacks)", file=sys.stderr)

    pin_path = args.out_dir / "arapahoe-pin-to-tag.json"
    pin_payload = {
        "snapshot": snapshot,
        "pinDigits": 9,
        "byPin": pin_map,
    }
    pin_path.write_text(json.dumps(pin_payload, separators=sep), encoding="utf-8")
    mb = pin_path.stat().st_size / (1024 * 1024)
    print(f"Wrote {pin_path} ({len(pin_map)} pins, {mb:.2f} MiB)", file=sys.stderr)

    situs_snapshot = {
        "bundledAsOf": bundled_as_of,
        "source": "Arapahoe County datamart: Main Parcel situs fields (Pin, SA*)",
        "taxYear": tax_year or None,
        "lookupNote": (
            "Keys match county address search rules: street number (+ optional range/suffix), "
            "street name without directionals (N,S,E,W,...) or types (St,Ave,...); optional unit."
        ),
    }
    situs_path = args.out_dir / "arapahoe-situs-to-pins.json"
    situs_path.write_text(
        json.dumps(
            {
                "snapshot": situs_snapshot,
                "lookupVersion": 1,
                "entryCount": len(situs_map),
                "byKey": situs_map,
            },
            separators=sep,
        ),
        encoding="utf-8",
    )
    sm = situs_path.stat().st_size / (1024 * 1024)
    print(
        f"Wrote {situs_path} ({len(situs_map)} keys, {sm:.2f} MiB)",
        file=sys.stderr,
    )

    parcel_snapshot = {
        "bundledAsOf": bundled_as_of,
        "source": (
            "Arapahoe County datamart: Main Parcel + Mart_DescrHeader + Mart_LegalParty "
            "+ Mart_RDE_LndAll + Mart_RDE_BLD + Mart_RDE_Xfob + Mart_Transfers + Mart_RDE_Permit "
            "+ State Class Codes xlsx "
            f"(lazy load after levy; sharded by {PARCEL_RECORD_SHARD_PREFIX_LEN}-digit PIN prefix)"
        ),
        "taxYear": tax_year or None,
    }
    write_parcel_record_shards(
        args.out_dir,
        parcel_record_map,
        parcel_snapshot,
        separators=sep,
    )


if __name__ == "__main__":
    main()
